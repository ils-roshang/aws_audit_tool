"""
modules/report_generator.py
-----------------------------
Phase 8: Generates PDF and Excel dashboard reports from the collected data.

PDF Structure (reportlab Platypus):
  1. Cover page
  2. Executive Summary
  3. Active Resources inventory
  4. Billing & Cost analysis
  5. Cost Optimisation recommendations
  6. Unused & Idle Resources
  7. Performance Analysis
  8. Security Audit

Excel Structure (openpyxl):
  - Summary sheet
  - Billing sheet (with embedded chart)
  - Resources sheet (per service type)
  - Recommendations sheet
  - Security Findings sheet
  - AI Insights sheet
  - AI Remediation Priority sheet
  - Performance Metrics sheet
  - Unused Resources sheet
"""

import io
import logging
import os
from datetime import datetime

# ── matplotlib (non-interactive backend MUST be set before any other import) ──
import matplotlib
matplotlib.use("Agg")

from utils.helpers import (
    time_series_chart,
    bar_chart,
    pie_chart,
    severity_donut,
    billing_trend_chart,
)

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════
# ██  P D F   G E N E R A T O R
# ═══════════════════════════════════════════════════════════════════════════

def generate_pdf(data: dict, output_path: str) -> None:
    """
    Generate a polished, client-facing PDF audit report.

    Sections:
        0. Cover page        — professional light theme, account info, KPI cards
        1. Table of Contents — section index
        2. Executive Summary — AI narrative + severity donut chart
        3. Billing & Cost    — period summary, MTD breakdown, 30-day services
        4. Resources         — service distribution + per-service detail tables
        5. Cost Optimisation — recommendations table + AI guidance
        6. Unused & Idle     — unused/idle resources table
        7. Performance       — spike alerts, AI root-cause, metric tables + charts
        8. Security Audit    — AI priority plan + colour-coded findings table
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        BaseDocTemplate, PageTemplate, Frame,
        Paragraph, Spacer, Table, TableStyle,
        PageBreak, CondPageBreak, Image, KeepTogether, HRFlowable,
    )

    # ── Safe section break ────────────────────────────────────────────────────
    # A4 content frame usable height is approximately 745 pt.
    # CondPageBreak(N) fires only when availHeight < N, i.e. when we are NOT
    # already at the very top of a fresh page.  Using 720 pt (≈ 96 % of 745)
    # means the break is skipped only when ≤ 25 pt (≈ 9 mm) of space has been
    # consumed since the last page advance — which cannot happen for any real
    # section content (the smallest section header alone is ≈ 60 pt).
    # This prevents blank pages that arise when a section's last KeepTogether
    # or table is pushed to a new page by ReportLab's flow engine and a
    # subsequent unconditional PageBreak() would produce an empty page.
    _SAFE_PB = CondPageBreak(720)

    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    PAGE_W, PAGE_H = A4
    MARGIN    = 1.8 * cm
    CONTENT_W = PAGE_W - 2 * MARGIN

    # ── Premium colour palette ───────────────────────────────────────────────
    NAVY        = colors.HexColor("#0F1C2E")   # deep navy — hero band & headers
    NAVY_MID    = colors.HexColor("#1A2D47")   # slightly lighter navy — accent bar
    BLUE        = colors.HexColor("#2563EB")   # vibrant blue — headings & accents
    BLUE_LIGHT  = colors.HexColor("#DBEAFE")   # pale blue — info stripe background
    TEAL        = colors.HexColor("#0EA5E9")   # sky teal — KPI accent rule
    SLATE       = colors.HexColor("#334155")   # body text
    MUTED       = colors.HexColor("#64748B")   # captions / footnotes
    ALT         = colors.HexColor("#F8FAFC")   # alternating row tint
    BORDER      = colors.HexColor("#CBD5E1")   # table grid
    WHITE       = colors.white
    GHOST       = colors.HexColor("#F1F5F9")   # subtle cell background
    # Severity colours
    HIGH_BG     = colors.HexColor("#FEF2F2")
    HIGH_FG     = colors.HexColor("#B91C1C")
    MED_BG      = colors.HexColor("#FFFBEB")
    MED_FG      = colors.HexColor("#B45309")
    LOW_BG      = colors.HexColor("#F0FDF4")
    LOW_FG      = colors.HexColor("#15803D")
    INFO_BG     = colors.HexColor("#EFF6FF")
    INFO_FG     = colors.HexColor("#1D4ED8")
    # KPI card accent colours
    KPI_SPEND   = colors.HexColor("#1E40AF")   # deep blue — spend
    KPI_SAVE    = colors.HexColor("#065F46")   # emerald — savings
    KPI_RES     = colors.HexColor("#1E3A5F")   # slate blue — resources
    KPI_REC     = colors.HexColor("#7C2D12")   # rust — recommendations

    # ── Style factory ────────────────────────────────────────────────────────
    _base = getSampleStyleSheet()

    _ps_cache: dict = {}
    def _ps(name, **kw):
        """Create and cache a ParagraphStyle to avoid duplicate-name warnings."""
        key = name + str(sorted(kw.items()))
        if key not in _ps_cache:
            _ps_cache[key] = ParagraphStyle(name, parent=_base["Normal"], **kw)
        return _ps_cache[key]

    S = {
        # ── Cover ──
        "hero_tag":  _ps("hero_tag",  fontName="Helvetica",      fontSize=8,
                         textColor=colors.HexColor("#93C5FD"),
                         alignment=TA_CENTER, spaceAfter=6),
        "cov_title": _ps("cov_title", fontName="Helvetica-Bold", fontSize=24,
                         textColor=WHITE, alignment=TA_CENTER, leading=30),
        "cov_sub":   _ps("cov_sub",   fontName="Helvetica",      fontSize=10,
                         textColor=colors.HexColor("#94A3B8"),
                         alignment=TA_CENTER, spaceAfter=4),
        # ── KPI cards ──
        "kpi_val":   _ps("kpi_val",   fontName="Helvetica-Bold", fontSize=16,
                         textColor=NAVY, alignment=TA_CENTER),
        "kpi_lbl":   _ps("kpi_lbl",   fontName="Helvetica",      fontSize=7,
                         textColor=MUTED, alignment=TA_CENTER),
        # ── Section / headings ──
        "sec_hdr":   _ps("sec_hdr",   fontName="Helvetica-Bold", fontSize=10,
                         textColor=WHITE, leading=15),
        "sec_sub":   _ps("sec_sub",   fontName="Helvetica",      fontSize=7.5,
                         textColor=colors.HexColor("#BAE6FD"), leading=11),
        "h2":        _ps("h2",        fontName="Helvetica-Bold", fontSize=9,
                         textColor=BLUE,
                         spaceBefore=8, spaceAfter=3, leading=13),
        "h3":        _ps("h3",        fontName="Helvetica-Bold", fontSize=8,
                         textColor=SLATE, spaceBefore=5, spaceAfter=3),
        # ── Body text ──
        "body":      _ps("body",      fontName="Helvetica",      fontSize=8,
                         textColor=SLATE, leading=12, spaceAfter=4),
        "bold":      _ps("bold",      fontName="Helvetica-Bold", fontSize=8,
                         textColor=SLATE, leading=12),
        "caption":   _ps("caption",   fontName="Helvetica-Oblique", fontSize=7,
                         textColor=MUTED, alignment=TA_CENTER, spaceAfter=4),
        "disc":      _ps("disc",      fontName="Helvetica-Oblique", fontSize=6.5,
                         textColor=MUTED),
        "note":      _ps("note",      fontName="Helvetica-Oblique", fontSize=7,
                         textColor=MUTED, spaceAfter=4),
        # ── Tables ──
        "tbl_h":     _ps("tbl_h",     fontName="Helvetica-Bold", fontSize=7,
                         textColor=WHITE, leading=10),
        "tbl_c":     _ps("tbl_c",     fontName="Helvetica",      fontSize=7,
                         textColor=SLATE, leading=10),
        "tbl_c_r":   _ps("tbl_c_r",   fontName="Helvetica",      fontSize=7,
                         textColor=SLATE, leading=10, alignment=TA_RIGHT),
        # ── Severity pill ──
        "sev_h":     _ps("sev_h",     fontName="Helvetica-Bold", fontSize=7,
                         textColor=HIGH_FG),
        "sev_m":     _ps("sev_m",     fontName="Helvetica-Bold", fontSize=7,
                         textColor=MED_FG),
        "sev_l":     _ps("sev_l",     fontName="Helvetica-Bold", fontSize=7,
                         textColor=LOW_FG),
        "sev_i":     _ps("sev_i",     fontName="Helvetica-Bold", fontSize=7,
                         textColor=INFO_FG),
        # ── Advice / callout box ──
        "adv":       _ps("adv",       fontName="Helvetica-Oblique", fontSize=7.5,
                         textColor=MED_FG, backColor=MED_BG,
                         leftIndent=8, rightIndent=8, spaceAfter=5, leading=11),
        "bullet":    _ps("bullet",    fontName="Helvetica",      fontSize=8,
                         textColor=SLATE, leftIndent=14, leading=12, spaceAfter=2),
        # ── ToC ──
        "toc_title": _ps("toc_title", fontName="Helvetica-Bold", fontSize=17,
                         textColor=NAVY, spaceAfter=10),
        "toc_item":  _ps("toc_item",  fontName="Helvetica",      fontSize=8.5,
                         textColor=SLATE, leading=15, leftIndent=6),
        "toc_num":   _ps("toc_num",   fontName="Helvetica",      fontSize=8.5,
                         textColor=MUTED, leading=15, alignment=TA_RIGHT),
    }

    # ── Running page header / footer  ────────────────────────────────────────
    def _on_page(canv, doc):
        """Drawn on every interior page (cover is page 1, skipped)."""
        if doc.page == 1:
            return
        canv.saveState()

        # ── Header bar ──
        # Solid navy strip across full page width
        canv.setFillColor(NAVY)
        canv.rect(0, PAGE_H - 14 * mm, PAGE_W, 14 * mm, fill=1, stroke=0)
        # Teal accent line under strip
        canv.setStrokeColor(TEAL)
        canv.setLineWidth(1.5)
        canv.line(0, PAGE_H - 14 * mm, PAGE_W, PAGE_H - 14 * mm)
        # Report name (left)
        canv.setFont("Helvetica-Bold", 7.5)
        canv.setFillColor(WHITE)
        canv.drawString(MARGIN, PAGE_H - 9 * mm, "AWS INFRASTRUCTURE AUDIT REPORT")
        # Account info (right)
        acc_name = str(data.get("account_name", "N/A"))[:32]
        canv.setFont("Helvetica", 7)
        canv.setFillColor(colors.HexColor("#93C5FD"))
        canv.drawRightString(PAGE_W - MARGIN, PAGE_H - 9 * mm, acc_name)

        # ── Footer bar ──
        canv.setFillColor(GHOST)
        canv.rect(0, 0, PAGE_W, 10 * mm, fill=1, stroke=0)
        canv.setStrokeColor(BORDER)
        canv.setLineWidth(0.4)
        canv.line(0, 10 * mm, PAGE_W, 10 * mm)
        canv.setFont("Helvetica-Oblique", 6.5)
        canv.setFillColor(MUTED)
        gen_time_local = data.get("generated_at", "")[:19].replace("T", " ")
        canv.drawString(MARGIN, 3.5 * mm,
                        f"CONFIDENTIAL  ·  Generated {gen_time_local} UTC")
        canv.setFont("Helvetica-Bold", 7)
        canv.setFillColor(NAVY)
        canv.drawRightString(PAGE_W - MARGIN, 3.5 * mm, f"Page {doc.page}")

        canv.restoreState()

    # ── Cover page template (no header/footer) ───────────────────────────────
    cover_frame = Frame(
        MARGIN, MARGIN, CONTENT_W, PAGE_H - 2 * MARGIN,
        id="cover", leftPadding=0, rightPadding=0,
        topPadding=0, bottomPadding=0,
    )
    # Interior page frame — shorter top/bottom to leave room for the h/f bars
    page_frame = Frame(
        MARGIN, 14 * mm, CONTENT_W, PAGE_H - 14 * mm - 12 * mm,
        id="main", leftPadding=0, rightPadding=0,
        topPadding=4 * mm, bottomPadding=4 * mm,
    )

    def _on_cover(canv, doc):
        """Professional light-theme cover page for page 1."""
        canv.saveState()
        # Clean white base
        canv.setFillColor(WHITE)
        canv.rect(0, 0, PAGE_W, PAGE_H, fill=1, stroke=0)
        # Deep navy hero band — top 38% of page
        canv.setFillColor(NAVY)
        canv.rect(0, PAGE_H * 0.62, PAGE_W, PAGE_H * 0.38, fill=1, stroke=0)
        # 5 mm teal accent stripe at very top
        canv.setFillColor(TEAL)
        canv.rect(0, PAGE_H - 5 * mm, PAGE_W, 5 * mm, fill=1, stroke=0)
        # 2pt teal rule at band/body boundary
        canv.setStrokeColor(TEAL)
        canv.setLineWidth(2)
        canv.line(0, PAGE_H * 0.62, PAGE_W, PAGE_H * 0.62)
        # Light ghost band for KPI cards (middle 20%)
        canv.setFillColor(colors.HexColor("#F8FAFC"))
        canv.rect(0, PAGE_H * 0.26, PAGE_W, PAGE_H * 0.20, fill=1, stroke=0)
        canv.setStrokeColor(colors.HexColor("#E2E8F0"))
        canv.setLineWidth(0.5)
        canv.line(0, PAGE_H * 0.26, PAGE_W, PAGE_H * 0.26)
        canv.line(0, PAGE_H * 0.46, PAGE_W, PAGE_H * 0.46)
        # Light footer strip (bottom 10%)
        canv.setFillColor(colors.HexColor("#F1F5F9"))
        canv.rect(0, 0, PAGE_W, PAGE_H * 0.10, fill=1, stroke=0)
        canv.setStrokeColor(colors.HexColor("#CBD5E1"))
        canv.setLineWidth(0.5)
        canv.line(0, PAGE_H * 0.10, PAGE_W, PAGE_H * 0.10)
        canv.restoreState()

    doc = BaseDocTemplate(
        output_path, pagesize=A4,
        leftMargin=MARGIN,  rightMargin=MARGIN,
        topMargin=MARGIN,   bottomMargin=MARGIN,
        title="AWS Infrastructure Audit Report",
        author="AWS Audit Tool",
    )
    doc.addPageTemplates([
        PageTemplate(id="cover", frames=[cover_frame], onPage=_on_cover),
        PageTemplate(id="main",  frames=[page_frame],  onPage=_on_page),
    ])
    story = []

    # ── Helper: switch template ───────────────────────────────────────────────
    from reportlab.platypus import NextPageTemplate

    # ── Shared helpers ────────────────────────────────────────────────────────
    def _sec(title, subtitle=""):
        """
        Full-width section header: deep-navy title row + optional teal-text
        subtitle row.  A 3 pt teal top-border creates a sharp visual tag.
        """
        rows       = [[Paragraph(f"<b>{title}</b>", S["sec_hdr"])]]
        style_cmds = [
            ("BACKGROUND",    (0, 0), (-1, 0), NAVY),
            ("TOPPADDING",    (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("LEFTPADDING",   (0, 0), (-1, -1), 14),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 12),
            # Teal accent on the top edge of the header bar
            ("LINEABOVE",     (0, 0), (-1, 0), 3, TEAL),
        ]
        if subtitle:
            rows.append([Paragraph(subtitle, S["sec_sub"])])
            style_cmds += [
                ("BACKGROUND",    (0, 1), (-1, 1), NAVY_MID),
                ("TOPPADDING",    (0, 1), (-1, 1), 4),
                ("BOTTOMPADDING", (0, 1), (-1, 1), 4),
            ]
        t = Table(rows, colWidths=[CONTENT_W])
        t.setStyle(TableStyle(style_cmds))
        # Section headers always land at the top of their own page (forced by
        # PageBreak before each section), so no leading whitespace is needed.
        # The trailing spacer separates the header band from the first sub-section.
        return KeepTogether([t, Spacer(1, 4 * mm)])

    def _tbl(headers, rows, col_widths=None, sev_col=None, right_cols=None):
        """
        Styled data table.
        sev_col   — column index whose value drives per-row background tinting.
        right_cols — list of column indices to right-align.
        """
        right_cols = right_cols or []
        SEV_BG = {"HIGH": HIGH_BG, "MEDIUM": MED_BG, "LOW": LOW_BG, "INFO": INFO_BG}

        hdr_row = [Paragraph(str(h), S["tbl_h"]) for h in headers]
        tbl_data = [hdr_row]
        sev_cmds = []

        for i, row in enumerate(rows):
            cells = []
            for ci, c in enumerate(row):
                ps = S["tbl_c_r"] if ci in right_cols else S["tbl_c"]
                cells.append(Paragraph(str(c), ps))
            tbl_data.append(cells)
            if sev_col is not None:
                try:
                    bg = SEV_BG.get(str(rows[i][sev_col]).upper())
                    if bg:
                        sev_cmds.append(("BACKGROUND", (0, i + 1), (-1, i + 1), bg))
                except Exception:
                    pass

        t = Table(tbl_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT", spaceAfter=4)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [WHITE, ALT]),
            ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
            ("LINEBELOW",     (0, 0), (-1, 0),  1.0, TEAL),   # teal rule under header
            ("TOPPADDING",    (0, 0), (-1, 0),  5),
            ("BOTTOMPADDING", (0, 0), (-1, 0),  5),
            ("TOPPADDING",    (0, 1), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ] + sev_cmds))
        return t

    def _img(png_bytes, width=CONTENT_W, height=5.5 * cm):
        buf = io.BytesIO(png_bytes)
        return Image(buf, width=width, height=height)

    def _centered_img(png_bytes, width=8 * cm, height=8 * cm):
        """Embed a chart PNG centred horizontally within the content column."""
        buf = io.BytesIO(png_bytes)
        img = Image(buf, width=width, height=height)
        t   = Table([[img]], colWidths=[CONTENT_W])
        t.setStyle(TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")]))
        return t

    def _usd(v):
        try:
            return f"${float(v):,.2f}"
        except Exception:
            return "—"

    def _divider():
        """Thin horizontal rule for visual breathing room between sub-sections."""
        return HRFlowable(width=CONTENT_W, thickness=0.4, color=BORDER,
                          spaceAfter=3 * mm, spaceBefore=1 * mm)

    def _kpi_card(value, label, accent_color):
        """
        Single KPI card: coloured top accent bar + large value + small label.
        Returns a Table suitable for embedding in a multi-column row.
        """
        val_ps = _ps(f"kv_{label[:4]}", fontName="Helvetica-Bold", fontSize=18,
                     textColor=accent_color, alignment=TA_CENTER)
        lbl_ps = _ps(f"kl_{label[:4]}", fontName="Helvetica",      fontSize=7,
                     textColor=MUTED, alignment=TA_CENTER)
        inner = Table(
            [[Paragraph(value, val_ps)],
             [Paragraph(label, lbl_ps)]],
            colWidths=[CONTENT_W / 4 - 4 * mm],
        )
        inner.setStyle(TableStyle([
            ("LINEABOVE",     (0, 0), (-1, 0), 3, accent_color),
            ("BACKGROUND",    (0, 0), (-1, -1), WHITE),
            ("BOX",           (0, 0), (-1, -1), 0.4, BORDER),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ]))
        return inner

    # ── Pre-compute ──────────────────────────────────────────────────────────
    billing       = data.get("billing", {})
    resources     = data.get("resources", [])
    recs          = data.get("recommendations", [])
    sec_findings  = data.get("security_findings", [])
    ai_insights   = data.get("ai_insights", {})
    regions       = data.get("regions_scanned", [])
    performance   = data.get("performance", {})
    gen_time      = data.get("generated_at", "")[:19].replace("T", " ")
    billing_30d   = billing.get("30d", {}).get("total", 0.0)
    total_savings = sum(r.get("estimated_monthly_savings_usd", 0) for r in recs)
    high_c  = sum(1 for f in sec_findings if f.get("severity") == "HIGH")
    med_c   = sum(1 for f in sec_findings if f.get("severity") == "MEDIUM")
    low_c   = sum(1 for f in sec_findings if f.get("severity") == "LOW")
    info_c  = sum(1 for f in sec_findings if f.get("severity") == "INFO")

    # Performance helpers — build a flat per-resource summary from by_namespace.
    # Only include namespaces that are in METRICS_REGISTRY (user-created resources
    # such as EC2, RDS, Lambda, etc.).  Operational/service namespaces returned by
    # CloudWatch (Logs, Usage, Bedrock, Rekognition, Observability Admin, etc.) are
    # excluded — they are not user resources and add noise to the report.
    from registries.metrics_registry import METRICS_REGISTRY as _METRICS_REGISTRY
    _TRACKED_NAMESPACES = set(_METRICS_REGISTRY.keys())

    _by_ns_raw = performance.get("by_namespace", {})
    _by_ns     = {ns: v for ns, v in _by_ns_raw.items() if ns in _TRACKED_NAMESPACES}
    _spikes    = [
        s for s in performance.get("spike_correlation", [])
        if s.get("namespace") in _TRACKED_NAMESPACES
    ]
    _perf_window = "7d"   # default window shown in PDF

    def _perf_summary() -> list:
        """
        Build a flat list of per-resource metric summaries across registered namespaces.
        Each entry: {namespace, resource_id, region, metric, avg, min_v, max_v, unit}
        Uses the 7-day window; falls back to 30d if 7d has no datapoints.
        """
        rows = []
        for ns, windows in _by_ns.items():
            resource_list = windows.get(_perf_window) or windows.get("30d") or []
            for rec in resource_list:
                rid    = rec.get("resource_id", "—")
                region = rec.get("region", "—")
                for metric_name, win_data in rec.get("metrics", {}).items():
                    pts = win_data.get(_perf_window) or win_data.get("30d") or []
                    values = [dp["value"] for dp in pts if dp.get("value") is not None]
                    if not values:
                        continue
                    _raw_unit  = pts[0].get("unit", "") if pts else ""
                    # Convert Bytes → GB for storage/memory metrics
                    _divisor   = 1024 ** 3 if _raw_unit == "Bytes" else 1
                    _disp_unit = "GB" if _raw_unit == "Bytes" else _raw_unit
                    rows.append({
                        "namespace":   ns,
                        "resource_id": rid,
                        "region":      region,
                        "metric":      metric_name,
                        "avg":         round((sum(values) / len(values)) / _divisor, 4),
                        "min_v":       round(min(values) / _divisor, 4),
                        "max_v":       round(max(values) / _divisor, 4),
                        "unit":        _disp_unit,
                    })
        return rows

    _perf_rows = _perf_summary()

    # Section labels used in both ToC and headings
    _sections = [
        ("Executive Summary",          "2"),
        ("Billing & Cost Analysis",     "3"),
        ("Resources Inventory",         "4"),
        ("Cost Optimisation",           "5"),
        ("Security Audit",             "6"),
    ]

    # ════════════════════════════════════════════════════════════════════════
    # 0 ░░ COVER PAGE  (professional light theme — drawn by _on_cover callback)
    # ════════════════════════════════════════════════════════════════════════

    # Push content into the navy hero band (top 38% ≈ 113 mm from top of A4)
    story.append(Spacer(1, 32 * mm))

    # Eyebrow tag
    story.append(Paragraph(
        "AWS  CLOUD  INFRASTRUCTURE  AUDIT",
        _ps("cov_eyebrow", fontName="Helvetica", fontSize=8,
            textColor=colors.HexColor("#93C5FD"), alignment=TA_CENTER, spaceAfter=5),
    ))
    story.append(Spacer(1, 3 * mm))

    # Main title — white text on navy hero
    story.append(Paragraph(
        f"{data.get('account_name', 'Infrastructure')}<br/>"
        "<font size='20'>Audit Report</font>",
        _ps("cov_title2", fontName="Helvetica-Bold", fontSize=28,
            textColor=WHITE, alignment=TA_CENTER, leading=36),
    ))
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(
        "Cost · Performance · Security · Rightsizing",
        _ps("cov_sub2", fontName="Helvetica", fontSize=9,
            textColor=colors.HexColor("#94A3B8"), alignment=TA_CENTER),
    ))
    story.append(Spacer(1, 8 * mm))

    # Account meta strip — still inside navy hero band
    acc_ps = _ps("covinfo2", fontName="Helvetica", fontSize=8.5,
                 textColor=colors.HexColor("#94A3B8"), alignment=TA_CENTER)
    info_cov = Table([[
        Paragraph(f"<b>Account ID</b><br/>{data.get('account_id', 'N/A')}", acc_ps),
        Paragraph(f"<b>Prepared</b><br/>{gen_time} UTC",                    acc_ps),
        Paragraph(f"<b>Regions</b><br/>{len(regions)} scanned",             acc_ps),
        Paragraph(f"<b>Resources</b><br/>{len(resources)} inventoried",     acc_ps),
    ]], colWidths=[CONTENT_W / 4] * 4)
    info_cov.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), NAVY_MID),
        ("LINEABOVE",     (0, 0), (-1,  0), 0.5, TEAL),
        ("LINEBELOW",     (0, 0), (-1, -1), 0.5, TEAL),
        ("TOPPADDING",    (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(info_cov)

    # ── KPI cards — white cards on ghost band (middle section) ──────────────
    story.append(Spacer(1, 14 * mm))

    def _kpi_light(value, label, accent):
        """Light-theme KPI card with coloured top-border accent."""
        v_ps = _ps(f"kl_v_{label}", fontName="Helvetica-Bold", fontSize=20,
                   textColor=accent, alignment=TA_CENTER)
        l_ps = _ps(f"kl_l_{label}", fontName="Helvetica",      fontSize=7.5,
                   textColor=MUTED, alignment=TA_CENTER)
        inner = Table([[Paragraph(value, v_ps)], [Paragraph(label, l_ps)]],
                      colWidths=[CONTENT_W / 4 - 6])
        inner.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), WHITE),
            ("BOX",           (0, 0), (-1, -1), 0.8, accent),
            ("LINEABOVE",     (0, 0), (-1,  0), 3,   accent),
            ("TOPPADDING",    (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        return inner

    kpi_spend = _kpi_light(_usd(billing_30d),   "30-Day AWS Spend",            colors.HexColor("#1E40AF"))
    kpi_save  = _kpi_light(_usd(total_savings), "Potential Monthly Savings",   colors.HexColor("#065F46"))
    kpi_res   = _kpi_light(str(len(resources)), "Resources Inventoried",       colors.HexColor("#1E3A5F"))
    kpi_rec   = _kpi_light(str(len(recs)),      "Rightsizing Recommendations", colors.HexColor("#7C2D12"))
    kpi_row   = Table([[kpi_spend, kpi_save, kpi_res, kpi_rec]],
                      colWidths=[CONTENT_W / 4] * 4)
    kpi_row.setStyle(TableStyle([
        ("LEFTPADDING",  (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(kpi_row)
    story.append(Spacer(1, 10 * mm))

    # ── Security severity summary — light pastel cards ───────────────────────
    sev_lbl_ps = _ps("cov_slbl", fontName="Helvetica",      fontSize=7,
                     textColor=MUTED,   alignment=TA_CENTER)
    sev_val_h  = _ps("cov_svh",  fontName="Helvetica-Bold", fontSize=15,
                     textColor=HIGH_FG, alignment=TA_CENTER)
    sev_val_m  = _ps("cov_svm",  fontName="Helvetica-Bold", fontSize=15,
                     textColor=MED_FG,  alignment=TA_CENTER)
    sev_val_l  = _ps("cov_svl",  fontName="Helvetica-Bold", fontSize=15,
                     textColor=LOW_FG,  alignment=TA_CENTER)
    sev_val_i  = _ps("cov_svi",  fontName="Helvetica-Bold", fontSize=15,
                     textColor=INFO_FG, alignment=TA_CENTER)

    story.append(Paragraph(
        "Security Findings Summary",
        _ps("cov_sec_lbl", fontName="Helvetica-Bold", fontSize=8,
            textColor=SLATE, alignment=TA_CENTER, spaceAfter=3),
    ))
    sv_row = Table([
        [Paragraph(str(high_c),  sev_val_h), Paragraph(str(med_c),  sev_val_m),
         Paragraph(str(low_c),   sev_val_l), Paragraph(str(info_c), sev_val_i)],
        [Paragraph("HIGH",   sev_lbl_ps), Paragraph("MEDIUM", sev_lbl_ps),
         Paragraph("LOW",    sev_lbl_ps), Paragraph("INFO",   sev_lbl_ps)],
    ], colWidths=[CONTENT_W / 4] * 4)
    sv_row.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), HIGH_BG),
        ("BACKGROUND", (1, 0), (1, -1), MED_BG),
        ("BACKGROUND", (2, 0), (2, -1), LOW_BG),
        ("BACKGROUND", (3, 0), (3, -1), INFO_BG),
        ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, BORDER),
        ("LINEABOVE",     (0, 0), (-1,  0), 2,   TEAL),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(sv_row)
    story.append(Spacer(1, 8 * mm))

    # Regions list
    if regions:
        story.append(Paragraph(
            f"<font color='#64748B'>Regions scanned:  </font>"
            f"<font color='#334155'><b>{'  ·  '.join(regions)}</b></font>",
            _ps("cov_reg2", fontName="Helvetica", fontSize=7.5,
                textColor=SLATE, alignment=TA_CENTER),
        ))
        story.append(Spacer(1, 4 * mm))

    # Confidential footer
    story.append(HRFlowable(
        width=CONTENT_W, thickness=0.4, color=BORDER,
        spaceBefore=2 * mm, spaceAfter=4 * mm,
    ))
    story.append(Paragraph(
        "CONFIDENTIAL  ·  This report is prepared exclusively for authorised recipients. "
        "Distribution to third parties is strictly prohibited.",
        _ps("cov_disc2", fontName="Helvetica-Oblique", fontSize=7,
            textColor=MUTED, alignment=TA_CENTER),
    ))

    # Switch to interior template starting from next page
    story.append(NextPageTemplate("main"))
    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════════════
    # 1 ░░ TABLE OF CONTENTS
    # ════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("Contents", S["toc_title"]))
    story.append(HRFlowable(width=CONTENT_W, thickness=2, color=TEAL,
                            spaceBefore=0, spaceAfter=6 * mm))

    toc_entries = [
        ("1.", "Executive Summary"),
        ("2.", "Billing &amp; Cost Analysis"),
        ("3.", "Resources Inventory"),
        ("4.", "Cost Optimisation Recommendations"),
        ("5.", "Unused &amp; Idle Resources"),
        ("6.", "Performance Analysis"),
        ("7.", "Trends"),
        ("8.", "Security Audit"),
    ]
    for num, label in toc_entries:
        toc_row = Table(
            [[Paragraph(f"{num}  {label}", S["toc_item"]),
              Paragraph("···", S["toc_num"])]],
            colWidths=[CONTENT_W * 0.85, CONTENT_W * 0.15],
        )
        toc_row.setStyle(TableStyle([
            ("LINEBELOW",    (0, 0), (-1, -1), 0.3, BORDER),
            ("TOPPADDING",   (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ("LEFTPADDING",  (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(toc_row)

    story.append(Spacer(1, 10 * mm))

    # Audit scope quick-facts box
    scope_ps  = _ps("scope_v",  fontName="Helvetica-Bold", fontSize=9,
                    textColor=NAVY, alignment=TA_CENTER)
    scope_lps = _ps("scope_l",  fontName="Helvetica",      fontSize=7,
                    textColor=MUTED, alignment=TA_CENTER)
    scope_tbl = Table([
        [Paragraph(str(len(resources)), scope_ps),
         Paragraph(str(len(recs)),      scope_ps),
         Paragraph(str(len(sec_findings)), scope_ps),
         Paragraph(str(len(regions)),   scope_ps)],
        [Paragraph("Resources",        scope_lps),
         Paragraph("Recommendations",  scope_lps),
         Paragraph("Security Findings",scope_lps),
         Paragraph("Regions",          scope_lps)],
    ], colWidths=[CONTENT_W / 4] * 4)
    scope_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), BLUE_LIGHT),
        ("BOX",           (0, 0), (-1, -1), 0.5, TEAL),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, BORDER),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(scope_tbl)
    # Each major section starts on its own page
    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════════════
    # 2 ░░ EXECUTIVE SUMMARY
    # ════════════════════════════════════════════════════════════════════════
    story.append(_sec("1.  Executive Summary"))

    exec_text = ai_insights.get("executive_summary", "")
    if exec_text:
        story.append(Paragraph(exec_text, S["body"]))
    else:
        story.append(Paragraph(
            f"This report covers <b>{len(resources)}</b> active AWS resources across "
            f"<b>{len(regions)}</b> region(s). "
            f"Total spend over the past 30 days was <b>{_usd(billing_30d)}</b>. "
            f"<b>{len(recs)}</b> right-sizing recommendations were identified with combined "
            f"potential monthly savings of <b>{_usd(total_savings)}</b>. "
            f"The security audit identified <b>{high_c}</b> HIGH, <b>{med_c}</b> MEDIUM, "
            f"and <b>{low_c}</b> LOW severity findings.",
            S["body"],
        ))

    # Key metrics recap strip
    story.append(Spacer(1, 3 * mm))
    metrics_ps_v = _ps("exec_kv", fontName="Helvetica-Bold", fontSize=11,
                        textColor=NAVY, alignment=TA_CENTER)
    metrics_ps_l = _ps("exec_kl", fontName="Helvetica",      fontSize=7,
                        textColor=MUTED, alignment=TA_CENTER)
    metrics = [
        (_usd(billing_30d),   "30-Day Spend"),
        (_usd(total_savings), "Potential Savings/mo"),
        (str(high_c),         "HIGH Findings"),
        (str(len(recs)),      "Recommendations"),
    ]
    m_row_vals = [Paragraph(v, metrics_ps_v) for v, _ in metrics]
    m_row_lbls = [Paragraph(l, metrics_ps_l) for _, l in metrics]
    m_tbl = Table([m_row_vals, m_row_lbls], colWidths=[CONTENT_W / 4] * 4)
    m_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), GHOST),
        ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, BORDER),
        ("LINEABOVE",     (0, 0), (-1, 0),  2,   TEAL),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(m_tbl)

    if high_c + med_c + low_c > 0:
        story.append(Spacer(1, 3 * mm))
        try:
            donut_png = severity_donut(high_c, med_c, low_c, "Security Findings by Severity")
            story.append(KeepTogether([
                _centered_img(donut_png, width=8 * cm, height=8 * cm),
                Paragraph("Figure 1 — Security findings distribution by severity", S["caption"]),
            ]))
        except Exception as e:
            logger.debug(f"Severity donut failed: {e}")

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════════════
    # 3 ░░ BILLING & COST ANALYSIS
    # ════════════════════════════════════════════════════════════════════════
    cur_month  = billing.get("current_month", {})
    cm_total   = cur_month.get("total", 0.0)
    story.append(_sec(
        "2.  Billing &amp; Cost Analysis",
        f"30-Day Total: {_usd(billing_30d)}   ·   Current Month: {_usd(cm_total)}",
    ))

    # --- Cost Period Summary ---
    story.append(Paragraph("Cost Period Summary", S["h2"]))
    period_rows = []
    for win_key, label in [("7d", "Last 7 Days"), ("15d", "Last 15 Days"), ("30d", "Last 30 Days")]:
        bw   = billing.get(win_key, {})
        tot  = bw.get("total", 0.0)
        svcs = bw.get("by_service", [])
        # Use actual days_elapsed stored by cost_analyzer; fall back to the
        # window key integer only when data is unavailable (e.g. dry-run).
        days = bw.get("days_elapsed") or int(win_key[:-1])
        top  = svcs[0] if svcs else {}
        p    = bw.get("period", {})
        period_rows.append([
            label,
            f"{p.get('start','')}  →  {p.get('end','')}",
            _usd(tot),
            _usd(tot / days) if days else "—",
            str(len(svcs)),
            top.get("service", "—"),
        ])
    story.append(_tbl(
        ["Period", "Date Range", "Total Cost", "Daily Avg", "Services", "Top Service"],
        period_rows,
        col_widths=[2.2 * cm, 4.5 * cm, 2.6 * cm, 2.0 * cm, 1.8 * cm,
                    CONTENT_W - 13.1 * cm],
        right_cols=[2, 3],
    ))
    story.append(Spacer(1, 3 * mm))

    # --- Current Month Breakdown ---
    cm_name  = cur_month.get("month_name",    "Current Month")
    cm_svcs  = cur_month.get("by_service",    [])
    cm_days  = cur_month.get("days_elapsed",  0)
    cm_dim   = cur_month.get("days_in_month", 30)
    cm_rem   = cur_month.get("days_remaining",0)
    cm_per   = cur_month.get("period", {})
    story.append(Paragraph(
        f"{cm_name} — Month-to-Date Billing  "
        f"({cm_days} of {cm_dim} days completed, {cm_rem} remaining)",
        S["h2"],
    ))
    if cm_svcs:
        cm_rows = []
        for idx, svc in enumerate(cm_svcs, 1):
            cost = svc.get("cost", 0)
            pct  = (cost / cm_total * 100) if cm_total else 0
            cm_rows.append([
                idx,
                svc.get("service", "—"),
                _usd(cost),
                f"{pct:.1f}%",
                _usd(cost / cm_days) if cm_days else "—",
            ])
        story.append(_tbl(
            ["#", "AWS Service", f"{cm_name} Cost", "% of Total", f"Daily Avg ({cm_days}d)"],
            cm_rows,
            col_widths=[1 * cm, 6.5 * cm, 3 * cm, 2.5 * cm, 3 * cm],
            right_cols=[2, 3, 4],
        ))
        story.append(Paragraph(
            f"Month-to-date total: <b>{_usd(cm_total)}</b>  ·  "
            f"Period: {cm_per.get('start','')} to {cm_per.get('end','')}",
            S["note"],
        ))
    else:
        story.append(Paragraph(
            "No finalised billing data available for the current month yet.", S["body"],
        ))
    story.append(Spacer(1, 3 * mm))

    # --- 30-Day Service Breakdown ---
    story.append(Paragraph("30-Day Service Cost Breakdown (All Services)", S["h2"]))
    _b30       = billing.get("30d", {})
    top_svcs   = _b30.get("by_service", [])
    _30d_days  = _b30.get("days_elapsed") or 30  # actual days in the 30d window
    if top_svcs:
        svc_rows = [
            [i, s["service"], _usd(s["cost"]),
             f"{s['cost'] / billing_30d * 100:.1f}%" if billing_30d else "—",
             _usd(s["cost"] / _30d_days) if _30d_days else "—"]
            for i, s in enumerate(top_svcs, 1)
        ]
        story.append(_tbl(
            ["#", "AWS Service", "30-Day Cost", "% of Total", "Daily Avg"],
            svc_rows,
            col_widths=[1 * cm, 6.5 * cm, 3 * cm, 2.5 * cm, 3 * cm],
            right_cols=[2, 3, 4],
        ))
    else:
        story.append(Paragraph("No billing data available.", S["body"]))

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════════════
    # 4 ░░ RESOURCES INVENTORY
    # ════════════════════════════════════════════════════════════════════════
    story.append(_sec(
        "3.  Resources Inventory",
        f"{len(resources)} active resources across {len(regions)} region(s)",
    ))

    # Service distribution summary table
    svc_counts: dict = {}
    for r in resources:
        svc_counts[r.get("service", "unknown").upper()] = (
            svc_counts.get(r.get("service", "unknown").upper(), 0) + 1
        )
    if svc_counts:
        svc_sum = [
            [i, svc, cnt, f"{cnt / len(resources) * 100:.1f}%"]
            for i, (svc, cnt) in enumerate(
                sorted(svc_counts.items(), key=lambda x: -x[1]), 1
            )
        ]
        story.append(Paragraph("Resources by Service", S["h2"]))
        story.append(_tbl(
            ["#", "Service", "Count", "% of Total"],
            svc_sum,
            col_widths=[1 * cm, 8 * cm, 3.5 * cm, 3.5 * cm],
        ))
        story.append(Spacer(1, 3 * mm))

    # Per-service detail tables
    services_grouped: dict = {}
    for r in resources:
        services_grouped.setdefault(r.get("service", "unknown"), []).append(r)

    # ── Service-specific table column definitions ────────────────────────────
    def _svc_tbl_def(svc: str, r_list: list):
        """
        Return (headers, rows, col_widths) with columns tailored to the AWS
        service so the table always shows the most meaningful attributes
        instead of a generic Type / Engine / Status layout.

        Any new service is handled by the generic fallback automatically.
        To add a bespoke layout, add an elif branch below — no other code
        changes are needed.
        """
        row_limit = 9999  # No cap: show every resource in the PDF

        def _m(r):
            """Shorthand metadata accessor."""
            return r.get("metadata", {})

        if svc == "ec2":
            headers    = ["Instance Name", "Region", "Instance Type", "State", "Public IP"]
            col_widths = [CONTENT_W * 0.35, CONTENT_W * 0.16, CONTENT_W * 0.20,
                          CONTENT_W * 0.15, CONTENT_W * 0.14]
            rows = [
                [r.get("name", "—")[:40], r.get("region", "—"),
                 _m(r).get("instance_type", "—"),
                 _m(r).get("state", "—"),
                 _m(r).get("public_ip", "") or "Private"]
                for r in r_list[:row_limit]
            ]

        elif svc == "rds":
            headers    = ["DB Identifier", "Region", "Engine / Version",
                          "Instance Class", "Multi-AZ", "Status"]
            col_widths = [CONTENT_W * 0.25, CONTENT_W * 0.13, CONTENT_W * 0.22,
                          CONTENT_W * 0.18, CONTENT_W * 0.10, CONTENT_W * 0.12]
            rows = [
                [r.get("name", "—")[:30], r.get("region", "—"),
                 (f"{_m(r).get('engine','?')} {_m(r).get('engine_version','')}").strip()[:24],
                 _m(r).get("instance_class", "—"),
                 "Yes" if _m(r).get("multi_az") else "No",
                 _m(r).get("status", "—")]
                for r in r_list[:row_limit]
            ]

        elif svc == "lambda":
            headers    = ["Function Name", "Region", "Runtime", "Memory (MB)", "Timeout (s)"]
            col_widths = [CONTENT_W * 0.38, CONTENT_W * 0.16, CONTENT_W * 0.18,
                          CONTENT_W * 0.14, CONTENT_W * 0.14]
            rows = [
                [r.get("name", "—")[:44], r.get("region", "—"),
                 _m(r).get("runtime", "—"),
                 str(_m(r).get("memory_mb", "—")),
                 str(_m(r).get("timeout_sec", "—"))]
                for r in r_list[:row_limit]
            ]

        elif svc == "s3":
            headers    = ["Bucket Name", "Region", "Versioning", "Public Access"]
            col_widths = [CONTENT_W * 0.44, CONTENT_W * 0.18, CONTENT_W * 0.18,
                          CONTENT_W * 0.20]
            rows = [
                [r.get("name", "—")[:52],
                 _m(r).get("bucket_region", r.get("region", "—")),
                 _m(r).get("versioning", "Disabled") or "Disabled",
                 "Blocked" if _m(r).get("public_access_blocked") else "OPEN (!)"]
                for r in r_list[:row_limit]
            ]

        elif svc == "dynamodb":
            headers    = ["Table Name", "Region", "Billing Mode",
                          "Item Count", "Encryption", "Status"]
            col_widths = [CONTENT_W * 0.28, CONTENT_W * 0.13, CONTENT_W * 0.16,
                          CONTENT_W * 0.13, CONTENT_W * 0.16, CONTENT_W * 0.14]
            rows = [
                [r.get("name", "—")[:32], r.get("region", "—"),
                 _m(r).get("billing_mode", "PROVISIONED"),
                 f"{int(_m(r).get('item_count', 0) or 0):,}",
                 _m(r).get("encryption_type", "DISABLED"),
                 _m(r).get("status", "—")]
                for r in r_list[:row_limit]
            ]

        elif svc == "elasticache":
            headers    = ["Cluster ID", "Region", "Engine / Version",
                          "Node Type", "Nodes", "Status"]
            col_widths = [CONTENT_W * 0.23, CONTENT_W * 0.13, CONTENT_W * 0.20,
                          CONTENT_W * 0.20, CONTENT_W * 0.10, CONTENT_W * 0.14]
            rows = [
                [r.get("name", "—")[:28], r.get("region", "—"),
                 (f"{_m(r).get('engine','?')} {_m(r).get('engine_version','')}").strip()[:22],
                 _m(r).get("node_type", "—"),
                 str(_m(r).get("num_nodes", "—")),
                 _m(r).get("status", "—")]
                for r in r_list[:row_limit]
            ]

        elif svc == "ecs":
            headers    = ["Service Name", "Region", "Cluster",
                          "Running / Desired", "Launch Type"]
            col_widths = [CONTENT_W * 0.28, CONTENT_W * 0.14, CONTENT_W * 0.22,
                          CONTENT_W * 0.20, CONTENT_W * 0.16]
            rows = [
                [r.get("name", "—")[:32], r.get("region", "—"),
                 (_m(r).get("cluster", "—") or "—")[-28:],
                 f"{_m(r).get('running_count','?')} / {_m(r).get('desired_count','?')}",
                 _m(r).get("launch_type", "FARGATE") or "FARGATE"]
                for r in r_list[:row_limit]
            ]

        elif svc == "eks":
            headers    = ["Cluster Name", "Region", "K8s Version",
                          "Public Endpoint", "Status"]
            col_widths = [CONTENT_W * 0.30, CONTENT_W * 0.16, CONTENT_W * 0.16,
                          CONTENT_W * 0.20, CONTENT_W * 0.18]
            rows = [
                [r.get("name", "—")[:35], r.get("region", "—"),
                 _m(r).get("version", "—"),
                 "Yes" if _m(r).get("endpoint_public_access") else "Private",
                 _m(r).get("status", "—")]
                for r in r_list[:row_limit]
            ]

        elif svc == "redshift":
            headers    = ["Cluster ID", "Region", "Node Type",
                          "Nodes", "Encrypted", "Status"]
            col_widths = [CONTENT_W * 0.25, CONTENT_W * 0.13, CONTENT_W * 0.20,
                          CONTENT_W * 0.10, CONTENT_W * 0.14, CONTENT_W * 0.18]
            rows = [
                [r.get("name", "—")[:28], r.get("region", "—"),
                 _m(r).get("node_type", "—"),
                 str(_m(r).get("num_nodes", "—")),
                 "Yes" if _m(r).get("encrypted") else "No",
                 _m(r).get("status", "—")]
                for r in r_list[:row_limit]
            ]

        elif svc == "sqs":
            headers    = ["Queue Name", "Region", "Messages",
                          "Retention (days)", "KMS Encrypted"]
            col_widths = [CONTENT_W * 0.34, CONTENT_W * 0.16, CONTENT_W * 0.13,
                          CONTENT_W * 0.21, CONTENT_W * 0.16]
            rows = []
            for r in r_list[:row_limit]:
                ret_sec = _m(r).get("message_retention_sec", 0) or 0
                try:
                    ret_days = str(int(ret_sec) // 86400)
                except (ValueError, TypeError):
                    ret_days = str(ret_sec)
                rows.append([
                    r.get("name", "—")[:40], r.get("region", "—"),
                    str(_m(r).get("approx_messages", "0") or "0"),
                    ret_days,
                    "Yes" if _m(r).get("kms_key") else "No",
                ])

        elif svc == "vpc":
            headers    = ["VPC Name / ID", "Region", "CIDR Block",
                          "State", "Default?", "Subnets"]
            col_widths = [CONTENT_W * 0.25, CONTENT_W * 0.13, CONTENT_W * 0.18,
                          CONTENT_W * 0.12, CONTENT_W * 0.12, CONTENT_W * 0.20]
            rows = [
                [r.get("name", "—")[:28], r.get("region", "—"),
                 _m(r).get("cidr_block", "—"),
                 _m(r).get("state", "—"),
                 "Yes" if _m(r).get("is_default") else "No",
                 str(_m(r).get("subnet_count", "—"))]
                for r in r_list[:row_limit]
            ]

        elif svc == "subnet":
            headers    = ["Subnet Name / ID", "Region", "CIDR Block",
                          "AZ", "VPC ID", "Public IP?"]
            col_widths = [CONTENT_W * 0.25, CONTENT_W * 0.12, CONTENT_W * 0.16,
                          CONTENT_W * 0.14, CONTENT_W * 0.18, CONTENT_W * 0.15]
            rows = [
                [r.get("name", "—")[:28], r.get("region", "—"),
                 _m(r).get("cidr_block", "—"),
                 _m(r).get("availability_zone", "—"),
                 _m(r).get("vpc_id", "—"),
                 "Yes" if _m(r).get("auto_assign_public_ip") else "No"]
                for r in r_list[:row_limit]
            ]

        elif svc == "igw":
            headers    = ["Gateway Name / ID", "Region", "Attached VPC", "State"]
            col_widths = [CONTENT_W * 0.35, CONTENT_W * 0.16, CONTENT_W * 0.30, CONTENT_W * 0.19]
            rows = [
                [r.get("name", "—")[:40], r.get("region", "—"),
                 _m(r).get("attached_vpc", "Detached"),
                 _m(r).get("state", "—")]
                for r in r_list[:row_limit]
            ]

        elif svc == "nat-gateway":
            headers    = ["NAT GW Name / ID", "Region", "Type", "State", "Public IP", "VPC"]
            col_widths = [CONTENT_W * 0.24, CONTENT_W * 0.12, CONTENT_W * 0.10,
                          CONTENT_W * 0.12, CONTENT_W * 0.16, CONTENT_W * 0.26]
            rows = [
                [r.get("name", "—")[:28], r.get("region", "—"),
                 _m(r).get("nat_type", "public").capitalize(),
                 _m(r).get("state", "—"),
                 _m(r).get("public_ip", "—"),
                 _m(r).get("vpc_id", "—")]
                for r in r_list[:row_limit]
            ]

        elif svc == "ebs":
            headers    = ["Volume Name / ID", "Region", "Type",
                          "Size (GB)", "State", "Attached To"]
            col_widths = [CONTENT_W * 0.25, CONTENT_W * 0.12, CONTENT_W * 0.10,
                          CONTENT_W * 0.12, CONTENT_W * 0.12, CONTENT_W * 0.29]
            rows = [
                [r.get("name", "—")[:28], r.get("region", "—"),
                 _m(r).get("volume_type", "—"),
                 str(_m(r).get("size_gb", "—")),
                 _m(r).get("state", "—"),
                 _m(r).get("attached_to", "") or "Unattached"]
                for r in r_list[:row_limit]
            ]

        elif svc == "ebs-snapshot":
            headers    = ["Snapshot Name / ID", "Region", "Volume",
                          "Size (GB)", "Encrypted", "State"]
            col_widths = [CONTENT_W * 0.26, CONTENT_W * 0.12, CONTENT_W * 0.20,
                          CONTENT_W * 0.12, CONTENT_W * 0.14, CONTENT_W * 0.16]
            rows = [
                [r.get("name", "—")[:30], r.get("region", "—"),
                 _m(r).get("volume_id", "—"),
                 str(_m(r).get("volume_size_gb", "—")),
                 "Yes" if _m(r).get("encrypted") else "No",
                 _m(r).get("state", "—")]
                for r in r_list[:row_limit]
            ]

        elif svc == "security-group":
            headers    = ["SG Name", "Region", "VPC", "Description",
                          "Inbound", "Outbound"]
            col_widths = [CONTENT_W * 0.18, CONTENT_W * 0.10, CONTENT_W * 0.14,
                          CONTENT_W * 0.47, CONTENT_W * 0.055, CONTENT_W * 0.055]
            rows = [
                [r.get("name", "—")[:24], r.get("region", "—"),
                 _m(r).get("vpc_id", "—"),
                 _m(r).get("description", "—"),
                 str(_m(r).get("inbound_rules", 0)),
                 str(_m(r).get("outbound_rules", 0))]
                for r in r_list[:row_limit]
            ]

        elif svc == "eip":
            headers    = ["Name / Alloc ID", "Region", "Public IP",
                          "Domain", "State", "Attached Instance"]
            col_widths = [CONTENT_W * 0.22, CONTENT_W * 0.12, CONTENT_W * 0.16,
                          CONTENT_W * 0.10, CONTENT_W * 0.16, CONTENT_W * 0.24]
            rows = [
                [r.get("name", "—")[:26], r.get("region", "—"),
                 _m(r).get("public_ip", "—"),
                 _m(r).get("domain", "vpc"),
                 _m(r).get("state", "—"),
                 _m(r).get("instance_id", "") or "Unassociated"]
                for r in r_list[:row_limit]
            ]

        elif svc == "loadbalancer":
            headers    = ["LB Name", "Region", "Type", "Scheme",
                          "State", "DNS Name"]
            col_widths = [CONTENT_W * 0.20, CONTENT_W * 0.12, CONTENT_W * 0.10,
                          CONTENT_W * 0.12, CONTENT_W * 0.10, CONTENT_W * 0.36]
            rows = [
                [r.get("name", "—")[:24], r.get("region", "—"),
                 (_m(r).get("lb_type", "application") or "application").capitalize(),
                 _m(r).get("scheme", "—"),
                 _m(r).get("state", "—"),
                 _m(r).get("dns_name", "—")[:42]]
                for r in r_list[:row_limit]
            ]

        elif svc == "sns":
            headers    = ["Topic Name", "Region", "Confirmed Subscriptions",
                          "Pending Subscriptions", "Encrypted"]
            col_widths = [CONTENT_W * 0.32, CONTENT_W * 0.14, CONTENT_W * 0.22,
                          CONTENT_W * 0.20, CONTENT_W * 0.12]
            rows = [
                [r.get("name", "—")[:38], r.get("region", "—"),
                 str(_m(r).get("subscriptions_confirmed", "0")),
                 str(_m(r).get("subscriptions_pending", "0")),
                 "Yes" if _m(r).get("kms_key") else "No"]
                for r in r_list[:row_limit]
            ]

        elif svc == "secretsmanager":
            headers    = ["Secret Name", "Region", "Auto Rotation",
                          "Last Rotated", "Last Accessed"]
            col_widths = [CONTENT_W * 0.32, CONTENT_W * 0.13, CONTENT_W * 0.15,
                          CONTENT_W * 0.20, CONTENT_W * 0.20]
            rows = [
                [r.get("name", "—")[:38], r.get("region", "—"),
                 "Enabled" if _m(r).get("rotation_enabled") else "Disabled",
                 str(_m(r).get("last_rotated", "—"))[:20],
                 str(_m(r).get("last_accessed", "—"))[:20]]
                for r in r_list[:row_limit]
            ]

        elif svc == "ecr":
            headers    = ["Repository Name", "Region", "Images",
                          "Scan on Push", "Encryption"]
            col_widths = [CONTENT_W * 0.35, CONTENT_W * 0.14, CONTENT_W * 0.12,
                          CONTENT_W * 0.18, CONTENT_W * 0.21]
            rows = [
                [r.get("name", "—")[:40], r.get("region", "—"),
                 str(_m(r).get("image_count", 0)),
                 "Yes" if _m(r).get("scan_on_push") else "No",
                 _m(r).get("encryption", "AES256")]
                for r in r_list[:row_limit]
            ]

        elif svc == "acm":
            headers    = ["Domain", "Region", "Status",
                          "Expiry", "Key Algorithm"]
            col_widths = [CONTENT_W * 0.30, CONTENT_W * 0.13, CONTENT_W * 0.15,
                          CONTENT_W * 0.22, CONTENT_W * 0.20]
            rows = [
                [r.get("name", "—")[:36], r.get("region", "—"),
                 _m(r).get("status", "—"),
                 str(_m(r).get("expiry", "—"))[:20],
                 _m(r).get("key_algorithm", "—")]
                for r in r_list[:row_limit]
            ]

        elif svc == "autoscaling":
            headers    = ["ASG Name", "Region", "Min / Max / Desired",
                          "Health Check", "AZs"]
            col_widths = [CONTENT_W * 0.30, CONTENT_W * 0.13, CONTENT_W * 0.22,
                          CONTENT_W * 0.17, CONTENT_W * 0.18]
            rows = [
                [r.get("name", "—")[:35], r.get("region", "—"),
                 f"{_m(r).get('min_size','?')} / {_m(r).get('max_size','?')} / {_m(r).get('desired_capacity','?')}",
                 _m(r).get("health_check_type", "EC2"),
                 str(_m(r).get("az_count", "—"))]
                for r in r_list[:row_limit]
            ]

        elif svc == "ssm":
            headers    = ["Document / Baseline Name", "Region", "Type",
                          "Format", "Platform", "Status"]
            col_widths = [CONTENT_W * 0.28, CONTENT_W * 0.12, CONTENT_W * 0.14,
                          CONTENT_W * 0.10, CONTENT_W * 0.18, CONTENT_W * 0.18]
            rows = [
                [r.get("name", "—")[:32], r.get("region", "—"),
                 _m(r).get("document_type", "—"),
                 _m(r).get("document_format", "—"),
                 _m(r).get("platform_types", "—")[:20],
                 _m(r).get("status", "—")]
                for r in r_list[:row_limit]
            ]

        elif svc == "cloudtrail":
            headers    = ["Trail Name", "Region", "Multi-Region",
                          "Logging", "Log Validation", "S3 Bucket"]
            col_widths = [CONTENT_W * 0.22, CONTENT_W * 0.12, CONTENT_W * 0.13,
                          CONTENT_W * 0.10, CONTENT_W * 0.15, CONTENT_W * 0.28]
            rows = [
                [r.get("name", "—")[:26], r.get("region", "—"),
                 "Yes" if _m(r).get("is_multi_region") else "No",
                 "On" if _m(r).get("is_logging") else "Off",
                 "Yes" if _m(r).get("log_validation") else "No",
                 _m(r).get("s3_bucket", "—")[:32]]
                for r in r_list[:row_limit]
            ]

        elif svc == "apigateway":
            headers    = ["API Name", "Region", "API ID",
                          "Endpoint Type", "Execute API Disabled"]
            col_widths = [CONTENT_W * 0.28, CONTENT_W * 0.13, CONTENT_W * 0.16,
                          CONTENT_W * 0.22, CONTENT_W * 0.21]
            rows = [
                [r.get("name", "—")[:32], r.get("region", "—"),
                 _m(r).get("api_id", "—"),
                 _m(r).get("endpoint_types", "—"),
                 "Yes" if _m(r).get("disable_execute_api") else "No"]
                for r in r_list[:row_limit]
            ]

        elif svc == "kinesis":
            headers    = ["Stream Name", "Region", "Mode",
                          "Shards", "Retention (h)", "Encryption"]
            col_widths = [CONTENT_W * 0.28, CONTENT_W * 0.13, CONTENT_W * 0.14,
                          CONTENT_W * 0.10, CONTENT_W * 0.16, CONTENT_W * 0.19]
            rows = [
                [r.get("name", "—")[:32], r.get("region", "—"),
                 _m(r).get("stream_mode", "PROVISIONED"),
                 str(_m(r).get("shard_count", "—")),
                 str(_m(r).get("retention_hours", "—")),
                 _m(r).get("encryption_type", "NONE")]
                for r in r_list[:row_limit]
            ]

        elif svc == "eventbridge":
            headers    = ["Rule Name", "Region", "State",
                          "Schedule", "Event Pattern", "Bus"]
            col_widths = [CONTENT_W * 0.26, CONTENT_W * 0.12, CONTENT_W * 0.10,
                          CONTENT_W * 0.22, CONTENT_W * 0.14, CONTENT_W * 0.16]
            rows = [
                [r.get("name", "—")[:30], r.get("region", "—"),
                 _m(r).get("state", "—"),
                 (_m(r).get("schedule", "") or "—")[:24],
                 _m(r).get("event_pattern", "No"),
                 (_m(r).get("event_bus", "default") or "default")]
                for r in r_list[:row_limit]
            ]

        elif svc == "stepfunctions":
            headers    = ["State Machine Name", "Region", "Type",
                          "Status", "Logging", "X-Ray Tracing"]
            col_widths = [CONTENT_W * 0.28, CONTENT_W * 0.13, CONTENT_W * 0.12,
                          CONTENT_W * 0.13, CONTENT_W * 0.16, CONTENT_W * 0.18]
            rows = [
                [r.get("name", "—")[:32], r.get("region", "—"),
                 _m(r).get("type", "—"),
                 _m(r).get("status", "—"),
                 _m(r).get("logging", "OFF"),
                 "Yes" if _m(r).get("tracing") else "No"]
                for r in r_list[:row_limit]
            ]

        else:
            # Generic fallback — IAM, CloudFront, Route53, Step Functions, etc.
            headers    = ["Name / ID", "Region"]
            col_widths = [CONTENT_W * 0.72, CONTENT_W * 0.28]
            rows = [
                [r.get("name", "—")[:80], r.get("region", "—")]
                for r in r_list[:row_limit]
            ]

        return headers, rows, col_widths

    # ── Per-service detail tables ────────────────────────────────────────────
    for svc, svc_res in sorted(services_grouped.items(), key=lambda x: -len(x[1])):
        # The label + first portion of the table are kept together to prevent
        # an orphaned heading at the bottom of a page.  Large tables are allowed
        # to split across pages (allowSplitting=True by default in ReportLab).
        svc_label = Paragraph(
            f"<b>{svc.upper()}</b>"
            f"<font color='#64748B' size='8'>  —  {len(svc_res)} resource(s)</font>",
            S["h2"],
        )
        headers, rows, col_widths = _svc_tbl_def(svc, svc_res)
        tbl = _tbl(headers, rows, col_widths=col_widths)
        # KeepTogether only enough to avoid a dangling heading: label + spacer + table.
        # If the whole block exceeds the remaining frame height, ReportLab will
        # split it naturally row-by-row instead of pushing it entirely to the next page.
        story.append(KeepTogether([svc_label, Spacer(1, 2 * mm)]))
        story.append(tbl)
        story.append(Spacer(1, 3 * mm))

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════════════
    # 5 ░░ COST OPTIMISATION RECOMMENDATIONS
    # ════════════════════════════════════════════════════════════════════════
    total_cost_increase = sum(r.get("estimated_monthly_cost_increase_usd", 0) for r in recs)
    _rec_subtitle = f"{len(recs)} recommendations   ·   Savings: {_usd(total_savings)}/month"
    if total_cost_increase > 0:
        _rec_subtitle += f"   ·   Upsize investment: +{_usd(total_cost_increase)}/month"
    story.append(_sec("4.  Cost Optimisation Recommendations", _rec_subtitle))

    if not recs:
        story.append(Paragraph(
            "No rightsizing recommendations were identified.", S["body"],
        ))
    else:
        # Build AI guidance lookup — all implementation steps + validation per resource.
        ai_lookup: dict = {}
        for e in (ai_insights.get("recommendations_enhanced", [])
                  if isinstance(ai_insights.get("recommendations_enhanced"), list) else []):
            if isinstance(e, dict) and e.get("resource_id"):
                steps      = e.get("implementation_steps", [])
                validation = e.get("validation", "")
                if steps:
                    parts = [f"<b>{i + 1}.</b> {str(s).strip()}" for i, s in enumerate(steps)]
                    if validation:
                        parts.append(f"<i>✓ {validation.strip()}</i>")
                    ai_lookup[e["resource_id"]] = "<br/>".join(parts)

        rows = []
        for rec in recs[:40]:
            rid      = rec.get("resource_name", rec.get("resource_id", "—"))
            rec_type = rec.get("recommendation_type", "")
            savings  = rec.get("estimated_monthly_savings_usd") or 0.0
            cost_inc = rec.get("estimated_monthly_cost_increase_usd") or 0.0
            rec_cfg  = rec.get("recommended_config", "")
            # Monthly Impact:
            #   Over-provisioned (no smaller type) → "N/A"
            #   Over-provisioned with confirmed downsize → "$X.XX" (exact or $0.00 if not priced)
            #   Under-provisioned with cost data → "+$X.XX"
            #   Under-provisioned, no pricing data → "Review"
            #   Review type → "Review"
            if rec_type == "Over-provisioned":
                monthly_impact = _usd(savings)   # $0.00 when no smaller class exists
            elif rec_type == "Under-provisioned":
                monthly_impact = f"+{_usd(cost_inc)}" if cost_inc else "Review"
            elif rec_type == "Review":
                monthly_impact = "Review"
            else:
                monthly_impact = "—"
            rows.append([
                rid[:30],
                rec.get("service", "—").upper(),
                rec.get("region", "—"),
                rec_type or "—",
                rec.get("current_config",     "—"),
                rec.get("recommended_config", "—"),
                # Reason — why the resource is flagged (metric-based evidence)
                rec.get("reason", "—")[:120],
                monthly_impact,
                rec.get("severity", "LOW"),
            ])

        # Build type-colouring commands for the Type column (index 3)
        _TYPE_FG = {
            "Over-provisioned":  colors.HexColor("#854D0E"),   # amber text
            "Under-provisioned": colors.HexColor("#9A3412"),   # red-orange text
            "Review":            colors.HexColor("#1E3A5F"),   # navy text
            "No-downsize target": colors.HexColor("#374151"),  # grey text
        }
        _TYPE_BG = {
            "Over-provisioned":  colors.HexColor("#FEF3C7"),   # amber tint
            "Under-provisioned": colors.HexColor("#FEE2E2"),   # red tint
            "Review":            colors.HexColor("#EFF6FF"),   # blue tint
            "No-downsize target": colors.HexColor("#F3F4F6"),  # grey tint
        }
        type_cmds = []
        for _ri, row in enumerate(rows):
            _rt = row[3]  # Type is index 3
            _bg = _TYPE_BG.get(_rt)
            if _bg:
                type_cmds.append(("BACKGROUND", (3, _ri + 1), (3, _ri + 1), _bg))

        # Columns (9 total, fit CONTENT_W = 17.4 cm):
        #  0 Resource | 1 Service | 2 Region | 3 Type | 4 Current |
        #  5 Recommended | 6 Reason | 7 Monthly Impact | 8 Priority
        rec_tbl = _tbl(
            ["Resource", "Service", "Region", "Type", "Current",
             "Recommended", "Reason", "Impact/mo", "Priority"],
            rows,
            col_widths=[2.8 * cm, 1.4 * cm, 1.5 * cm, 2.3 * cm, 1.6 * cm,
                        2.2 * cm, 2.8 * cm, 1.6 * cm, 1.2 * cm],
            sev_col=8,      # Priority is now index 8
            right_cols=[7], # Monthly Impact is now index 7
        )
        if type_cmds:
            rec_tbl.setStyle(TableStyle(type_cmds))
        story.append(rec_tbl)

        # AI one-line guidance callout box
        guided = [
            (rec, ai_lookup.get(rec.get("resource_name", rec.get("resource_id", ""))))
            for rec in recs[:40]
        ]
        guided = [(r, g) for r, g in guided if g]
        if guided:
            story.append(Spacer(1, 3 * mm))
            story.append(Paragraph("AI Implementation Guidance", S["h2"]))
            story.append(_divider())
            # Render as a two-column table: resource name | action
            guidance_rows = []
            for rec, guidance in guided:
                rid = rec.get("resource_name", rec.get("resource_id", ""))
                guidance_rows.append([
                    Paragraph(f"<b>{rid[:35]}</b>", S["tbl_c"]),
                    Paragraph(guidance, S["tbl_c"]),
                ])
            g_tbl = Table(guidance_rows,
                          colWidths=[5.5 * cm, CONTENT_W - 5.5 * cm],
                          spaceAfter=4)
            g_tbl.setStyle(TableStyle([
                ("ROWBACKGROUNDS", (0, 0), (-1, -1), [WHITE, GHOST]),
                ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
                ("LINEABOVE",     (0, 0), (-1, 0),  1.5, TEAL),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING",   (0, 0), (-1, -1), 6),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
                ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ]))
            story.append(g_tbl)

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════════════
    # 5 ░░ UNUSED & IDLE RESOURCES
    # ════════════════════════════════════════════════════════════════════════
    # Human-readable service name map for the PDF section (mirrors Excel sheet 8)
    _SVC_NAME_PDF: dict = {
        "AWS::EC2::Instance":                             "Amazon EC2",
        "AWS::EC2::Volume":                               "Amazon EBS",
        "AWS::EC2::EIP":                                  "Elastic IP",
        "AWS::EC2::VPC":                                  "Amazon VPC",
        "AWS::EC2::NatGateway":                           "NAT Gateway",
        "AWS::RDS::DBInstance":                           "Amazon RDS",
        "AWS::RDS::DBCluster":                            "Amazon Aurora",
        "AWS::ElastiCache::CacheCluster":                 "Amazon ElastiCache",
        "AWS::ElastiCache::ReplicationGroup":             "ElastiCache (Replication Group)",
        "AWS::Redshift::Cluster":                         "Amazon Redshift",
        "AWS::Lambda::Function":                          "AWS Lambda",
        "AWS::ECS::Service":                              "Amazon ECS",
        "AWS::ECS::Cluster":                              "Amazon ECS (Cluster)",
        "AWS::ElasticLoadBalancingV2::LoadBalancer":       "ALB / NLB",
        "AWS::ElasticLoadBalancing::LoadBalancer":         "Classic ELB",
        "AWS::Kinesis::Stream":                           "Amazon Kinesis",
        "AWS::SQS::Queue":                                "Amazon SQS",
        "AWS::SNS::Topic":                                "Amazon SNS",
        "AWS::ApiGateway::RestApi":                       "API Gateway (REST)",
        "AWS::ApiGatewayV2::Api":                         "API Gateway (HTTP)",
        "AWS::SageMaker::NotebookInstance":               "SageMaker Notebook",
    }

    unused_findings = sorted(
        [f for f in sec_findings
         if f.get("category") in ("unused_resource", "idle_resource")],
        key=lambda f: f.get("estimated_monthly_cost_usd", 0),
        reverse=True,
    )

    # Count totals for the section subtitle
    idle_count   = sum(1 for f in unused_findings if f.get("category") == "idle_resource")
    unused_count = sum(1 for f in unused_findings if f.get("category") == "unused_resource")

    story.append(_sec(
        "5.  Unused &amp; Idle Resources",
        f"Total: {len(unused_findings)}   ·   "
        f"Unused: {unused_count}   ·   Idle: {idle_count}",
    ))

    if not unused_findings:
        story.append(Paragraph(
            "No unused or idle resources were detected.", S["body"],
        ))
    else:
        # ── Badge summary row ─────────────────────────────────────────────────
        unused_ps = _ps("ur_uval", fontName="Helvetica-Bold", fontSize=18,
                        textColor=colors.HexColor("#B45309"), alignment=TA_CENTER)
        idle_ps   = _ps("ur_ival", fontName="Helvetica-Bold", fontSize=18,
                        textColor=colors.HexColor("#1D4ED8"), alignment=TA_CENTER)
        lbl_ps    = _ps("ur_lbl",  fontName="Helvetica",      fontSize=7.5,
                        textColor=MUTED, alignment=TA_CENTER)

        monthly_waste = sum(
            f.get("estimated_monthly_cost_usd", 0) or 0
            for f in unused_findings
        )
        waste_ps = _ps("ur_wval", fontName="Helvetica-Bold", fontSize=18,
                       textColor=colors.HexColor("#991B1B"), alignment=TA_CENTER)

        badge_row = Table([[
            Table([[Paragraph(str(unused_count), unused_ps)],
                   [Paragraph("Unused Resources", lbl_ps)]],
                  colWidths=[CONTENT_W / 3 - 4]),
            Table([[Paragraph(str(idle_count), idle_ps)],
                   [Paragraph("Idle Resources", lbl_ps)]],
                  colWidths=[CONTENT_W / 3 - 4]),
            Table([[Paragraph(_usd(monthly_waste), waste_ps)],
                   [Paragraph("Est. Monthly Waste", lbl_ps)]],
                  colWidths=[CONTENT_W / 3 - 4]),
        ]], colWidths=[CONTENT_W / 3] * 3)
        badge_row.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#FEF3C7")),
            ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#EFF6FF")),
            ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#FEE2E2")),
            ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, BORDER),
            ("LINEABOVE",     (0, 0), (-1,  0), 2.5, TEAL),
            ("LEFTPADDING",   (0, 0), (-1, -1), 3),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(badge_row)
        story.append(Spacer(1, 3 * mm))

        # ── Findings detail table ─────────────────────────────────────────────
        story.append(Paragraph("Unused / Idle Resources Detail", S["h2"]))
        story.append(_divider())
        ur_rows = []
        for f in unused_findings:
            cat      = f.get("category", "unused_resource")
            status   = "Idle" if cat == "idle_resource" else "Unused"
            rtype    = f.get("resource_type", "")
            svc_name = _SVC_NAME_PDF.get(rtype, rtype)
            cost_val = f.get("estimated_monthly_cost_usd")
            cost_str = _usd(cost_val) if (cost_val is not None and cost_val > 0) else "$0.00"
            ur_rows.append([
                status,
                svc_name,
                f.get("region",        "—"),
                f.get("resource_id",   "—")[:35],
                f.get("issue",         "—"),
                cost_str,
            ])

        ur_tbl = _tbl(
            ["Status", "Service", "Region", "Resource", "Issue", "Est. Cost/mo"],
            ur_rows,
            col_widths=[1.8 * cm, 3.2 * cm, 2.6 * cm, 3.4 * cm,
                        CONTENT_W - 14.0 * cm, 2.5 * cm],
        )
        # Override Status cell colours (amber=Unused, blue=Idle)
        if ur_tbl and ur_rows:
            ur_tbl.setStyle(TableStyle([
                ("TEXTCOLOR", (0, r + 1), (0, r + 1),
                 colors.HexColor("#1D4ED8") if ur_rows[r][0] == "Idle"
                 else colors.HexColor("#B45309"))
                for r in range(len(ur_rows))
            ]))
        story.append(ur_tbl)
        story.append(Spacer(1, 4 * mm))

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════════════
    # 6 ░░ PERFORMANCE ANALYSIS
    # ════════════════════════════════════════════════════════════════════════
    _spike_count = len(_spikes)
    _ns_count    = len(_by_ns)
    story.append(_sec(
        "6.  Performance Analysis",
        f"Window: 7-day   ·   Namespaces: {_ns_count}   ·   Anomalies detected: {_spike_count}",
    ))

    if not _perf_rows and not _spikes:
        story.append(Paragraph(
            "No CloudWatch performance data was collected for this run. "
            "Ensure resources have CloudWatch metrics enabled and the "
            "<b>CW_PERIOD</b> and <b>TIME_WINDOWS</b> settings are configured correctly.",
            S["body"],
        ))
    else:
        # ── Spike / Anomaly Alerts ──────────────────────────────────────────
        if _spikes:
            story.append(Paragraph("Performance Anomalies", S["h2"]))
            story.append(_divider())
            # Severity badge row
            _sp_high = sum(1 for s in _spikes if s.get("severity") == "HIGH")
            _sp_med  = sum(1 for s in _spikes if s.get("severity") == "MEDIUM")
            sp_badge = Table([[
                Table([[Paragraph(str(_sp_high), _ps("sph", fontName="Helvetica-Bold", fontSize=18, textColor=HIGH_FG, alignment=TA_CENTER))],
                        [Paragraph("HIGH Anomalies", _ps("sphl", fontName="Helvetica", fontSize=7, textColor=HIGH_FG, alignment=TA_CENTER))]],
                       colWidths=[CONTENT_W / 2 - 4]),
                Table([[Paragraph(str(_sp_med), _ps("spm", fontName="Helvetica-Bold", fontSize=18, textColor=MED_FG, alignment=TA_CENTER))],
                        [Paragraph("MEDIUM Anomalies", _ps("spml", fontName="Helvetica", fontSize=7, textColor=MED_FG, alignment=TA_CENTER))]],
                       colWidths=[CONTENT_W / 2 - 4]),
            ]], colWidths=[CONTENT_W / 2] * 2)
            sp_badge.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (0, -1), HIGH_BG),
                ("BACKGROUND",    (1, 0), (1, -1), MED_BG),
                ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
                ("INNERGRID",     (0, 0), (-1, -1), 0.4, BORDER),
                ("LINEABOVE",     (0, 0), (-1,  0), 2.5, TEAL),
                ("TOPPADDING",    (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING",   (0, 0), (-1, -1), 3),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
            ]))
            story.append(sp_badge)
            story.append(Spacer(1, 4 * mm))

            # Spike detail table
            spike_rows = [
                [
                    s.get("severity",    "—"),
                    s.get("namespace",   "—").replace("AWS/", ""),
                    s.get("resource_id", "—")[:35],
                    s.get("metric",      "—"),
                    s.get("window",      "—"),
                    f"{s.get('avg', 0):.2f}",
                    f"{s.get('z_score', 0):.2f}",
                ]
                for s in sorted(_spikes, key=lambda x: x.get("z_score", 0), reverse=True)
            ]
            story.append(_tbl(
                ["Severity", "Service", "Resource", "Metric", "Window", "Avg Value", "Z-Score"],
                spike_rows,
                col_widths=[1.8 * cm, 2.2 * cm, 4.2 * cm, 3.8 * cm,
                            1.5 * cm, 2.2 * cm, 1.9 * cm],
                sev_col=0,
                right_cols=[5, 6],
            ))
            story.append(Spacer(1, 3 * mm))

            # AI root-cause analysis
            _rca = ai_insights.get("root_cause_analysis", [])
            _valid_rca = [r for r in _rca if isinstance(r, dict) and r.get("resource_id")]
            if _valid_rca:
                story.append(Paragraph("AI Root-Cause Analysis", S["h2"]))
                story.append(_divider())
                rca_rows = []
                for r in _valid_rca:
                    rca_rows.append([
                        Paragraph(f"<b>{r.get('resource_id', '—')[:35]}</b>", S["tbl_c"]),
                        Paragraph(r.get("metric",       "—"), S["tbl_c"]),
                        Paragraph(r.get("likely_cause", "—"), S["tbl_c"]),
                        Paragraph(r.get("action",       "—"), S["tbl_c"]),
                    ])
                rca_tbl = Table(
                    rca_rows,
                    colWidths=[4.5 * cm, 3.0 * cm,
                               (CONTENT_W - 7.5 * cm) * 0.50,
                               (CONTENT_W - 7.5 * cm) * 0.50],
                    spaceAfter=4,
                )
                rca_tbl.setStyle(TableStyle([
                    ("ROWBACKGROUNDS", (0, 0), (-1, -1), [WHITE, GHOST]),
                    ("GRID",          (0, 0), (-1, -1), 0.3, BORDER),
                    ("LINEABOVE",     (0, 0), (-1, 0),  1.5, TEAL),
                    ("TOPPADDING",    (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("LEFTPADDING",   (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
                    ("VALIGN",        (0, 0), (-1, -1), "TOP"),
                ]))
                story.append(rca_tbl)
                story.append(Spacer(1, 3 * mm))

        # ── Per-namespace metric summary tables ─────────────────────────────
        if _perf_rows:
            story.append(Paragraph("CloudWatch Metric Summary — 7-Day Window", S["h2"]))
            story.append(_divider())

            # Group by namespace
            import collections as _coll
            _ns_grouped = _coll.defaultdict(list)
            for pr in _perf_rows:
                _ns_grouped[pr["namespace"]].append(pr)

            for ns, ns_recs in sorted(_ns_grouped.items()):
                # Namespace sub-heading
                _ns_res_count  = len(set(r["resource_id"] for r in ns_recs))
                _ns_met_count  = len(ns_recs)
                story.append(Paragraph(
                    f"<b>{ns.replace('AWS/', '')}</b>"
                    f"<font color='#64748B' size='7'>  \u2014  {_ns_res_count} resource(s)"
                    f"  \u00b7  {_ns_met_count} metric series</font>",
                    S["h2"],
                ))
                # Sort: resource_id, then metric
                ns_recs_sorted = sorted(ns_recs, key=lambda r: (r["resource_id"], r["metric"]))
                perf_tbl_rows = [
                    [
                        r["resource_id"][:32],
                        r["region"],
                        r["metric"],
                        r["unit"][:12] if r["unit"] else "—",
                        f"{r['avg']:.2f}",
                        f"{r['min_v']:.2f}",
                        f"{r['max_v']:.2f}",
                    ]
                    for r in ns_recs_sorted
                ]
                story.append(_tbl(
                    ["Resource", "Region", "Metric", "Unit", "Avg", "Min", "Max"],
                    perf_tbl_rows,
                    col_widths=[4.0 * cm, 2.2 * cm, 3.8 * cm, 1.8 * cm,
                                1.8 * cm, 1.8 * cm, 1.6 * cm],
                    right_cols=[4, 5, 6],
                ))
                story.append(Spacer(1, 3 * mm))

            # ── Time-series chart for top spiking resources ─────────────────
            if _spikes:
                story.append(Spacer(1, 2 * mm))
                story.append(Paragraph("Performance Trend Charts — Top Anomalies", S["h2"]))
                story.append(_divider())
                # Render up to 4 charts (one per top-spiking resource+metric)
                _chart_keys = []
                _seen_chart = set()
                for sp in _spikes:
                    _key = (sp["namespace"], sp["resource_id"], sp["metric"])
                    if _key not in _seen_chart:
                        _seen_chart.add(_key)
                        _chart_keys.append(sp)
                    if len(_chart_keys) >= 4:
                        break

                for sp in _chart_keys:
                    _ns  = sp["namespace"]
                    _rid = sp["resource_id"]
                    _met = sp["metric"]
                    # Collect all windows for this resource+metric
                    _series = []
                    _windows_data = _by_ns.get(_ns, {})
                    for _win_key in ["7d", "15d", "30d"]:
                        _res_list = _windows_data.get(_win_key, [])
                        for _res in _res_list:
                            if _res["resource_id"] != _rid:
                                continue
                            _pts = _res.get("metrics", {}).get(_met, {}).get(_win_key, [])
                            if _pts:
                                from datetime import datetime as _dt
                                _ts = [_dt.fromisoformat(p["timestamp"].replace("Z", "+00:00")) for p in _pts]
                                _vs = [p["value"] for p in _pts]
                                _series.append({"label": _win_key, "timestamps": _ts, "values": _vs})
                                break  # one series per window is enough
                    if not _series:
                        continue
                    try:
                        _unit = _spikes[0].get("unit", "") if _spikes else ""
                        chart_png = time_series_chart(
                            _series,
                            title=f"{_rid[:30]}  —  {_met}",
                            ylabel=_met,
                            width=12, height=3.5,
                        )
                        story.append(KeepTogether([
                            _centered_img(chart_png, width=CONTENT_W, height=3.8 * cm),
                            Paragraph(
                                f"Figure — {_met} trend for <b>{_rid}</b>"
                                f" (anomaly avg: {sp['avg']:.2f}, Z-score: {sp['z_score']:.2f})",
                                S["caption"],
                            ),
                        ]))
                        story.append(Spacer(1, 4 * mm))
                    except Exception as _ce:
                        logger.debug(f"Performance chart failed for {_rid}/{_met}: {_ce}")

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════════════
    # 7 ░░ TRENDS & PATTERNS
    # ════════════════════════════════════════════════════════════════════════
    _tp  = data.get("trends", {})
    _ct  = _tp.get("cost_trends",        {})
    _ptl = _tp.get("performance_trends", [])
    _fp  = _tp.get("fleet_patterns",     {})
    _sp  = _tp.get("security_patterns",  {})
    _cto = _ct.get("overall", {})

    # Direction colour helpers
    _DC = {"INCREASING": colors.HexColor("#B91C1C"),
            "DECREASING": colors.HexColor("#15803D"),
            "STABLE":     colors.HexColor("#1D4ED8")}
    _DB = {"INCREASING": colors.HexColor("#FEF2F2"),
            "DECREASING": colors.HexColor("#F0FDF4"),
            "STABLE":     colors.HexColor("#EFF6FF")}

    def _dir_arrow(d): return "▲" if d == "INCREASING" else ("▼" if d == "DECREASING" else "◆")

    # Pull the numbers we actually need
    _tc_dir   = _cto.get("direction", "STABLE")
    _tc_wow   = _cto.get("wow_delta_usd",     0)
    _tc_wow_p = _cto.get("wow_delta_pct",     0)
    # Use the calendar-aware projection (actual MTD ÷ elapsed × days_in_month) as the
    # estimated monthly bill; it is more stable than daily_7d × 30 which can spike.
    _tc_proj  = _cto.get("cm_projected", _cto.get("projected_monthly", 0))
    _fp_eff   = _fp.get("efficiency_score",    0)
    _fp_save  = _fp.get("total_savings_potential", 0)
    _sp_sev   = _sp.get("severity_distribution", {})
    _sp_high  = _sp_sev.get("HIGH", 0)
    _wow_s    = "+" if _tc_wow >= 0 else ""
    _dir_col  = _DC.get(_tc_dir, MUTED)
    _eff_col  = LOW_FG if _fp_eff >= 80 else (MED_FG if _fp_eff >= 50 else HIGH_FG)

    # Fleet health counts — needed by the glance cards (also used in section C)
    _fp_total = _fp.get("total_resources",       0)
    _fp_over  = _fp.get("overprovisioned_count", 0)
    _fp_under = _fp.get("underprovisioned_count",0)
    _fp_right = max(_fp_total - _fp_over - _fp_under, 0)

    # ── Month-over-month comparison ──────────────────────────────────────────
    # current-month daily rate vs 30-day daily rate (from trend_analyzer)
    _mom_pct     = _cto.get("mom_delta_pct", 0)
    _mom_s       = "+" if _mom_pct >= 0 else ""
    # USD equivalent: annualise the daily rate difference to a monthly amount
    # mom_delta_usd = (cm_daily_rate - daily_30d) × 30 — same baseline as mom_delta_pct,
    # so both the USD amount and the % on Card 1 are derived from the same current-month rate.
    _mom_usd_raw = _cto.get(
        "mom_delta_usd",
        round((_cto.get("daily_avg_7d", 0) - _cto.get("daily_avg_30d", 0)) * 30, 2),
    )
    _mom_usd_s   = "+" if _mom_usd_raw >= 0 else "-"
    _mom_usd_abs = abs(_mom_usd_raw)
    # Colour: amber/red when spending is up; green when down; blue when flat
    _mom_col  = (HIGH_FG if _mom_pct > 5
                 else (MED_FG if _mom_pct > 0
                       else (LOW_FG if _mom_pct < -5 else INFO_FG)))

    story.append(_sec(
        "7.  Trends",
        "A quick-read snapshot of cost direction, performance changes, fleet health, "
        "and security patterns.",
    ))

    # ── Pre-compute WoW card values ───────────────────────────────────────────
    _prior_daily    = _cto.get("prior_7d_daily_avg", 0)
    _cur_daily      = _cto.get("daily_avg_7d",        0)
    _wow_daily_diff = _cur_daily - _prior_daily
    _wow_daily_pct  = ((_wow_daily_diff / _prior_daily) * 100) if _prior_daily else 0.0
    _wow_d_sign     = "+" if _wow_daily_diff >= 0 else ""
    _wow_d_col      = (HIGH_FG if _wow_daily_pct > 5
                       else (MED_FG if _wow_daily_pct > 0
                             else (LOW_FG if _wow_daily_pct < -5 else INFO_FG)))
    _wow_d_bg       = (HIGH_BG if _wow_daily_pct > 5
                       else (MED_BG if _wow_daily_pct > 0
                             else (LOW_BG if _wow_daily_pct < -5 else INFO_BG)))

    # Current-month daily rate — primary value on card 1.
    # Computed as MTD total ÷ days elapsed, same formula used in the MTD billing table.
    _cm_total_v         = billing.get("current_month", {}).get("total", 0.0)
    _cm_days_elapsed    = billing.get("current_month", {}).get("days_elapsed",  0)
    _cm_days_in_month   = billing.get("current_month", {}).get("days_in_month", 31)
    _cm_daily_rate_c1   = (_cm_total_v / _cm_days_elapsed) if _cm_days_elapsed else 0.0
    # Last month's daily rate — the 30-day rolling daily average is a close proxy
    # because the current month is still in progress, so most of the 30d window
    # is last month's data.
    _lm_daily_rate_c1   = _cto.get("daily_avg_30d", 0)

    # ── 4-card at-a-glance strip ──────────────────────────────────────────────
    def _glance_card(value_para, label_text, bg_col, accent_col=None):
        """
        KPI summary card.  An optional accent_col draws a 3pt coloured rule
        across the very top of the card for quick visual scanning.
        """
        lbl_ps = _ps(
            f"gl_{label_text[:8]}",
            fontName="Helvetica", fontSize=6.5,
            textColor=MUTED, alignment=TA_CENTER, leading=9,
        )
        inner = Table(
            [[value_para],
             [Paragraph(label_text, lbl_ps)]],
            colWidths=[CONTENT_W / 4 - 4],
        )
        style_cmds = [
            ("BACKGROUND",    (0, 0), (-1, -1), bg_col),
            ("TOPPADDING",    (0, 0), (0,   0), 12),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 10),
            ("TOPPADDING",    (0, 1), (-1,  1), 3),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ]
        if accent_col:
            style_cmds.append(("LINEABOVE", (0, 0), (-1, 0), 3, accent_col))
        inner.setStyle(TableStyle(style_cmds))
        return inner

    # ── Card 1 — Current Month vs Last Month ──────────────────────────────────
    # Primary value: current-month daily rate (/day) — mirrors Card 2's format so
    # both week and month comparisons are immediately comparable by any reader.
    # Supporting line: last month's daily rate + % change (same pattern as Card 2).
    _c1 = _glance_card(
        Paragraph(
            f"<font color='{NAVY.hexval()}'><b>{_usd(_cm_daily_rate_c1)}/day</b></font><br/>"
            f"<font color='{_mom_col.hexval()}' size='8'>"
            f"<b>{_dir_arrow(_tc_dir)}&nbsp;{_mom_s}{_mom_pct:.1f}%"
            f" · Last: {_usd(_lm_daily_rate_c1)}/day</b></font>",
            _ps("c1v", fontName="Helvetica-Bold", fontSize=14,
                textColor=NAVY, alignment=TA_CENTER, leading=18),
        ),
        "Current Month vs Last Month",
        _DB.get(_tc_dir, GHOST),
        accent_col=_mom_col,
    )

    # ── Card 2 — This Week's Daily Spend Rate ────────────────────────────────
    # Primary value: actual current-week daily rate.
    # Supporting line: last week's rate + % change so the comparison is explicit.
    _c2 = _glance_card(
        Paragraph(
            f"<font color='{NAVY.hexval()}'><b>{_usd(_cur_daily)}/day</b></font><br/>"
            f"<font color='{_wow_d_col.hexval()}' size='8'>"
            f"<b>{_dir_arrow('INCREASING' if _wow_daily_diff >= 0 else 'DECREASING')}"
            f"&nbsp;{_wow_d_sign}{_wow_daily_pct:.1f}% · Last: {_usd(_prior_daily)}/day"
            f"</b></font>",
            _ps("c2v", fontName="Helvetica-Bold", fontSize=14,
                textColor=NAVY, alignment=TA_CENTER, leading=18),
        ),
        "This Week Daily Rate",
        _wow_d_bg,
        accent_col=_wow_d_col,
    )

    # ── Card 3 — Projected Month Total ───────────────────────────────────────
    # Primary value only — calendar-aware projected month-end total.
    # An invisible second line (same leading as Cards 1 & 2) keeps the card
    # the same height as its neighbours in the 4-card strip.
    # Using &nbsp; (non-breaking space) instead of a plain space ensures
    # ReportLab does not strip the second line during XML parsing, which
    # would otherwise collapse the paragraph and make the card shorter.
    _c3 = _glance_card(
        Paragraph(
            f"<font size='14'><b>{_usd(_tc_proj)}</b></font><br/>"
            f"<font size='8'>&nbsp;</font>",
            _ps("c3v", fontName="Helvetica-Bold", fontSize=14,
                textColor=NAVY, alignment=TA_CENTER, leading=18),
        ),
        "Projected Month Total",
        GHOST,
        accent_col=TEAL,
    )

    # ── Card 4 — Fleet Right-sizing Score ────────────────────────────────────
    _c4 = _glance_card(
        Paragraph(
            f"<font color='{_eff_col.hexval()}'><b>{_fp_eff:.0f}%</b></font><br/>"
            f"<font color='{_eff_col.hexval()}' size='9'>"
            f"<b>{_fp_right} / {_fp_total} correct</b></font>",
            _ps("c4v", fontName="Helvetica-Bold", fontSize=13,
                textColor=_eff_col, alignment=TA_CENTER, leading=17),
        ),
        "Fleet Right-sizing Score",
        GHOST,
        accent_col=_eff_col,
    )

    _cards = Table([[_c1, _c2, _c3, _c4]], colWidths=[CONTENT_W / 4] * 4)
    _cards.setStyle(TableStyle([
        ("BOX",        (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID",  (0, 0), (-1, -1), 0.3, BORDER),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(_cards)
    story.append(Spacer(1, 6 * mm))

    # ══════════════════════════════════════════════════════════════════════════
    # A  Cost — What is happening to your bill?
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph("A.  Cost — What is happening to your bill?", S["h2"]))
    story.append(_divider())

    # One-sentence plain-English summary
    _cv7  = _cto.get("daily_avg_7d",  0)
    _cv30 = _cto.get("daily_avg_30d", 0)
    _accel = _cto.get("spend_acceleration", 0)
    _conf  = _cto.get("projection_confidence", "LOW")
    # Phrase acceleration relative to the overall direction to avoid contradictions
    # (e.g. "going up and slowing down" → "going up, though the rate is easing")
    if _tc_dir == "INCREASING":
        _accel_txt = ("and accelerating"               if _accel > 5
                      else ("though the rate is easing" if _accel < -5
                            else "at a steady pace"))
    elif _tc_dir == "DECREASING":
        _accel_txt = ("and accelerating downward"       if _accel < -5
                      else ("though the rate is easing" if _accel > 5
                            else "at a steady pace"))
    else:
        _accel_txt = "at a steady pace"

    # Remaining days — used in the projection explanation sentence.
    _cm_remaining = int(_cto.get("cm_remaining", max(_cm_days_in_month - _cm_days_elapsed, 0)))

    if _tc_dir == "INCREASING":
        _summary_text = (
            f"Your AWS bill is <b>going up</b> {_accel_txt} — your daily spend this week "
            f"({_usd(_cv7)}/day) is higher than your 30-day average ({_usd(_cv30)}/day).  "
            f"With {_usd(_cm_total_v)} already spent this month and {_cm_remaining} days remaining "
            f"at {_usd(_cv7)}/day, you are on track to spend around <b>{_usd(_tc_proj)} this month</b>."
        )
    elif _tc_dir == "DECREASING":
        _summary_text = (
            f"Your AWS bill is <b>coming down</b> {_accel_txt} — your daily spend this week "
            f"({_usd(_cv7)}/day) is lower than your 30-day average ({_usd(_cv30)}/day).  "
            f"With {_usd(_cm_total_v)} already spent this month and {_cm_remaining} days remaining "
            f"at {_usd(_cv7)}/day, your estimated monthly bill is <b>{_usd(_tc_proj)}</b>."
        )
    else:
        _summary_text = (
            f"Your AWS bill is <b>stable</b> — your daily spend this week ({_usd(_cv7)}/day) "
            f"is in line with your 30-day average ({_usd(_cv30)}/day).  "
            f"With {_usd(_cm_total_v)} already spent and {_cm_remaining} days remaining, "
            f"estimated monthly bill: <b>{_usd(_tc_proj)}</b>."
        )

    _sum_tbl = Table([[Paragraph(_summary_text, S["body"])]], colWidths=[CONTENT_W])
    _sum_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), _DB.get(_tc_dir, GHOST)),
        ("LINEBEFORE",    (0, 0), ( 0, -1), 4, _DC.get(_tc_dir, TEAL)),
        ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING",   (0, 0), (-1, -1), 12),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
    ]))
    story.append(_sum_tbl)
    story.append(Spacer(1, 4 * mm))

    # Service cost table: week-over-week AND month-over-month daily averages
    _svc_data = _ct.get("by_service", [])
    if _svc_data:
        # Build a per-service current-month daily rate lookup from the same
        # billing.current_month.by_service data used by the MTD table above.
        # This ensures "Cur Month /day" matches "Daily Avg (Nd)" in the MTD table
        # exactly — both compute  cost / days_elapsed  for the same period.
        _cm_days_svc = billing.get("current_month", {}).get("days_elapsed", 0) or 0
        _cm_svc_map: dict = {
            s["service"]: float(s.get("cost", 0) or 0) / _cm_days_svc
            for s in billing.get("current_month", {}).get("by_service", [])
            if _cm_days_svc > 0 and s.get("service")
        }

        story.append(Paragraph(
            "Daily Cost by Service — Week-over-Week &amp; Month-over-Month", S["h3"],
        ))
        _sc_rows = []
        for _sv in _svc_data[:15]:
            _sv_name  = _sv.get("service", "—")
            _sv_dir   = _sv.get("direction", "STABLE")
            _sv_col   = _DC.get(_sv_dir, MUTED).hexval()
            # Week-over-week (7d daily vs prior-7d daily derived from 15d window)
            _d7       = _sv.get("daily_avg_7d",  0)
            _d15      = _sv.get("daily_avg_15d", 0)   # last-week proxy
            _d30      = _sv.get("daily_avg_30d", 0)   # 30-day rolling (last-month proxy)
            # Actual current-month daily rate: MTD cost ÷ days elapsed — same
            # formula used in the Month-to-Date Billing table so both tables agree.
            _cm_rate  = _cm_svc_map.get(_sv_name, 0.0)
            # WoW: last 7 days vs prior 7 days (days 8-15 extracted from 15d window)
            _wow_chg  = ((_d7 - _d15) / _d15 * 100) if _d15 else (100.0 if _d7 else 0.0)
            # MoM: actual current-month rate vs 30-day rolling baseline
            _mom_chg  = ((_cm_rate - _d30) / _d30 * 100) if _d30 else (100.0 if _cm_rate else 0.0)
            _wow_sign = "+" if _wow_chg >= 0 else ""
            _mom_sign = "+" if _mom_chg >= 0 else ""
            # WoW colour (red up, green down, blue flat)
            _wow_col  = ("#B91C1C" if _wow_chg > 5
                         else ("#15803D" if _wow_chg < -5 else "#1D4ED8"))
            # MoM colour — uses the same colour logic as WoW
            _mom_col_svc = ("#B91C1C" if _mom_chg > 5
                            else ("#15803D" if _mom_chg < -5 else "#1D4ED8"))
            _sc_rows.append([
                _sv_name,
                # Week column pair
                _usd(_d15),   # prev week daily (days 8-15 average)
                _usd(_d7),    # this week daily (last 7 days average)
                f"<font color='{_wow_col}'><b>{_wow_sign}{_wow_chg:.0f}%</b></font>",
                # Month column pair
                _usd(_d30),   # last month proxy (30-day rolling avg)
                _usd(_cm_rate),  # actual current-month daily avg (MTD ÷ elapsed)
                f"<font color='{_mom_col_svc}'><b>{_mom_sign}{_mom_chg:.0f}%</b></font>",
            ])
        # Column widths: Service | PrevWk | ThisWk | WoW% | LastMo | CurMo | MoM%
        _svc_col_w = [
            3.8 * cm, 2.4 * cm, 2.4 * cm, 1.5 * cm,
            2.4 * cm, 2.4 * cm, 1.5 * cm,
        ]
        # Remaining space guard — clamp to CONTENT_W
        _svc_col_w[-1] = max(CONTENT_W - sum(_svc_col_w[:-1]), 1.5 * cm)
        _svc_tbl = _tbl(
            ["Service",
             "Prev Week /day", "This Week /day", "WoW",
             "Last Month /day", "Cur Month /day", "MoM"],
            _sc_rows,
            col_widths=_svc_col_w,
            right_cols=[1, 2, 4, 5],
        )
        story.append(_svc_tbl)

    # Biggest movers callout
    _growing = _ct.get("fastest_growing", [])
    _saving  = _ct.get("biggest_savers",  [])
    if _growing or _saving:
        story.append(Spacer(1, 3 * mm))
        _mv_rows = []
        if _growing:
            _mv_rows.append([
                Paragraph("<font color='#B91C1C'><b>▲  Rising fastest</b></font>",
                          _ps("rl", fontName="Helvetica-Bold", fontSize=8, textColor=HIGH_FG)),
                Paragraph(",   ".join(_growing), S["body"]),
            ])
        if _saving:
            _mv_rows.append([
                Paragraph("<font color='#15803D'><b>▼  Falling most</b></font>",
                          _ps("fl", fontName="Helvetica-Bold", fontSize=8, textColor=LOW_FG)),
                Paragraph(",   ".join(_saving), S["body"]),
            ])
        _mv_tbl = Table(_mv_rows, colWidths=[4.0 * cm, CONTENT_W - 4.0 * cm])
        _mv_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), BLUE_LIGHT),
            ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
            ("LINEABOVE",     (0, 0), (-1, 0),  2,   TEAL),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(_mv_tbl)

    story.append(Spacer(1, 6 * mm))

    # ══════════════════════════════════════════════════════════════════════════
    # B  Performance — Are any resources showing warning signs?
    # ══════════════════════════════════════════════════════════════════════════
    _ptl_important = [t for t in _ptl if t.get("severity") in ("HIGH", "MEDIUM")]
    _ptl_low       = [t for t in _ptl if t.get("severity") == "LOW"]

    story.append(Paragraph(
        "B.  Performance — Are any resources showing warning signs?", S["h2"],
    ))
    story.append(_divider())

    if not _ptl_important and not _ptl_low:
        story.append(Paragraph(
            "No significant performance changes were detected in the current "
            "monitoring window. All tracked resources are operating within "
            "normal metric thresholds.",
            S["body"],
        ))
        story.append(Spacer(1, 6 * mm))
    else:
        if _ptl_important:
            # Alert banner for HIGH-severity items
            _hi_count = sum(1 for t in _ptl_important if t.get("severity") == "HIGH")
            if _hi_count:
                _alert = Table(
                    [[Paragraph(
                        f"<b>&#9888;  {_hi_count} resource(s) need immediate attention</b> — "
                        f"critical metric changes detected (see red rows below).",
                        S["body"],
                    )]],
                    colWidths=[CONTENT_W],
                )
                _alert.setStyle(TableStyle([
                    ("BACKGROUND",    (0, 0), (-1, -1), HIGH_BG),
                    ("LINEBEFORE",    (0, 0), ( 0, -1), 4, HIGH_FG),
                    ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
                    ("TOPPADDING",    (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                    ("LEFTPADDING",   (0, 0), (-1, -1), 10),
                ]))
                story.append(_alert)
                story.append(Spacer(1, 3 * mm))

        # Simplified 4-column table: Severity | Resource | What changed | What to do
        _pt_display = (_ptl_important + _ptl_low)[:25]
        _pt_simple_rows = []
        for _pt in _pt_display:
            _pt_dir  = _pt.get("direction", "STABLE")
            _pt_col  = _DC.get(_pt_dir, MUTED).hexval()
            _pt_chg  = _pt.get("change_pct", 0)
            _pt_sign = "+" if _pt_chg >= 0 else ""
            _ns_s    = _pt.get("namespace", "").replace("AWS/", "")
            _pt_simple_rows.append([
                _pt.get("severity", "LOW"),
                f"{_pt.get('resource_id','—')[:22]}  ({_ns_s})",
                f"<font color='{_pt_col}'><b>{_dir_arrow(_pt_dir)} "
                f"{_pt.get('metric','—')}  {_pt_sign}{_pt_chg:.0f}%</b></font>",
                _pt.get("interpretation", ""),
            ])
        story.append(_tbl(
            ["Severity", "Resource  (Service)", "What Changed", "What It Means"],
            _pt_simple_rows,
            col_widths=[1.6*cm, 5.0*cm, 4.5*cm, CONTENT_W - 11.1*cm],
            sev_col=0,
        ))
        if len(_ptl) > 25:
            story.append(Paragraph(
                f"<i>… {len(_ptl) - 25} more performance changes not shown.</i>",
                S["note"],
            ))
        story.append(Spacer(1, 6 * mm))

    # ══════════════════════════════════════════════════════════════════════════
    # C  Fleet Health — Are your resources the right size?
    # ══════════════════════════════════════════════════════════════════════════
    story.append(Paragraph(
        "C.  Fleet Health — Are your resources the right size?", S["h2"],
    ))
    story.append(_divider())

    # _fp_total, _fp_over, _fp_under, _fp_right already computed above the cards strip
    _fp_eff_v = _fp_eff   # alias — same value, keeps section C self-consistent
    _fp_sav_v = _fp.get("total_savings_potential", 0)
    _fp_eff_c = _eff_col  # alias

    # 3-number summary (too big / just right / too small)
    # Each inner table has a padding command to add readable space between the
    # large numeric value and the descriptive label below it.
    _fh_inner_style = TableStyle([
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),   # space below the number
        ("TOPPADDING",    (0, 1), (-1, 1), 4),   # space above the label
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
    ])
    _fh_card1 = Table([
        [Paragraph(f"<b>{_fp_over}</b>",
                   _ps("fh1v", fontName="Helvetica-Bold", fontSize=20,
                       textColor=MED_FG, alignment=TA_CENTER))],
        [Paragraph("Too Big (over-provisioned)",
                   _ps("fh1l", fontName="Helvetica", fontSize=7,
                       textColor=MUTED, alignment=TA_CENTER))],
    ], colWidths=[CONTENT_W / 3 - 4])
    _fh_card1.setStyle(_fh_inner_style)

    _fh_card2 = Table([
        [Paragraph(f"<b>{_fp_right}</b>",
                   _ps("fh2v", fontName="Helvetica-Bold", fontSize=20,
                       textColor=LOW_FG, alignment=TA_CENTER))],
        [Paragraph("Just Right",
                   _ps("fh2l", fontName="Helvetica", fontSize=7,
                       textColor=MUTED, alignment=TA_CENTER))],
    ], colWidths=[CONTENT_W / 3 - 4])
    _fh_card2.setStyle(_fh_inner_style)

    _fh_card3 = Table([
        [Paragraph(f"<b>{_fp_under}</b>",
                   _ps("fh3v", fontName="Helvetica-Bold", fontSize=20,
                       textColor=HIGH_FG, alignment=TA_CENTER))],
        [Paragraph("Too Small (under-provisioned)",
                   _ps("fh3l", fontName="Helvetica", fontSize=7,
                       textColor=MUTED, alignment=TA_CENTER))],
    ], colWidths=[CONTENT_W / 3 - 4])
    _fh_card3.setStyle(_fh_inner_style)

    _fh_strip = Table([[_fh_card1, _fh_card2, _fh_card3]],
                      colWidths=[CONTENT_W / 3] * 3)
    _fh_strip.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (0, -1), MED_BG),
        ("BACKGROUND",    (1, 0), (1, -1), LOW_BG),
        ("BACKGROUND",    (2, 0), (2, -1), HIGH_BG),
        ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, BORDER),
        ("LINEABOVE",     (0, 0), (-1,  0), 3,   TEAL),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(_fh_strip)
    story.append(Spacer(1, 3 * mm))

    # Score + savings one-liner
    _score_line = (
        f"Fleet right-sizing score: <font color='{_fp_eff_c.hexval()}'>"
        f"<b>{_fp_eff_v:.0f}%</b></font>  "
        f"({_fp_right} of {_fp_total} resources are correctly sized)  ·  "
        f"Estimated monthly savings if all over-provisioned resources are rightsized: "
        f"<b>{_usd(_fp_sav_v)}</b>."
    )
    story.append(Paragraph(_score_line, S["body"]))
    story.append(Spacer(1, 3 * mm))

    # Action items — each over/under issue as a plain bullet
    _sys_issues = _fp.get("systemic_issues", [])
    if _sys_issues:
        story.append(Paragraph("What needs attention:", S["h3"]))
        for _si in _sys_issues:
            story.append(Paragraph(f"&#8226;  {_si}", S["bullet"]))
        story.append(Spacer(1, 3 * mm))

    # Top services with rightsizing potential
    _svc_recs = _fp.get("services_with_most_recs", [])
    if _svc_recs:
        story.append(Paragraph("Services with the most rightsizing opportunities", S["h3"]))
        _sr_rows = [
            [s["service"],
             str(s["count"]),
             _usd(s["savings"]),
             s["dominant_type"]]
            for s in _svc_recs
        ]
        story.append(_tbl(
            ["Service", "Count", "Potential Saving", "Primary Issue"],
            _sr_rows,
            col_widths=[3.5*cm, 1.5*cm, 3.5*cm, CONTENT_W - 8.5*cm],
            right_cols=[2],
        ))

    story.append(Spacer(1, 6 * mm))

    # ══════════════════════════════════════════════════════════════════════════
    # D  Security — Any repeating problems?
    # ══════════════════════════════════════════════════════════════════════════
    _sp_pats = _sp.get("recurring_patterns",   [])
    _sp_hot  = _sp.get("services_by_findings", [])
    _sp_hd   = _sp.get("high_risk_domains",    [])

    if _sp_pats or _sp_hot or _sp_sev:
        story.append(Paragraph(
            "D.  Security — Any repeating problems?", S["h2"],
        ))
        story.append(_divider())

        # Severity count badges (4 cards)
        _sp_hi_c = _sp_sev.get("HIGH",   0)
        _sp_me_c = _sp_sev.get("MEDIUM", 0)
        _sp_lo_c = _sp_sev.get("LOW",    0)
        _sp_in_c = _sp_sev.get("INFO",   0)
        _sb = Table(
            [[Paragraph(f"<b>{_sp_hi_c}</b>", _ps("d1v", fontName="Helvetica-Bold", fontSize=18, textColor=HIGH_FG, alignment=TA_CENTER)),
              Paragraph(f"<b>{_sp_me_c}</b>", _ps("d2v", fontName="Helvetica-Bold", fontSize=18, textColor=MED_FG,  alignment=TA_CENTER)),
              Paragraph(f"<b>{_sp_lo_c}</b>", _ps("d3v", fontName="Helvetica-Bold", fontSize=18, textColor=LOW_FG,  alignment=TA_CENTER)),
              Paragraph(f"<b>{_sp_in_c}</b>", _ps("d4v", fontName="Helvetica-Bold", fontSize=18, textColor=INFO_FG, alignment=TA_CENTER))],
             [Paragraph("HIGH",   _ps("d1l", fontName="Helvetica", fontSize=7, textColor=HIGH_FG, alignment=TA_CENTER)),
              Paragraph("MEDIUM", _ps("d2l", fontName="Helvetica", fontSize=7, textColor=MED_FG,  alignment=TA_CENTER)),
              Paragraph("LOW",    _ps("d3l", fontName="Helvetica", fontSize=7, textColor=LOW_FG,  alignment=TA_CENTER)),
              Paragraph("INFO",   _ps("d4l", fontName="Helvetica", fontSize=7, textColor=INFO_FG, alignment=TA_CENTER))]],
            colWidths=[CONTENT_W / 4] * 4,
        )
        _sb.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (0, -1), HIGH_BG),
            ("BACKGROUND",    (1, 0), (1, -1), MED_BG),
            ("BACKGROUND",    (2, 0), (2, -1), LOW_BG),
            ("BACKGROUND",    (3, 0), (3, -1), INFO_BG),
            ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, BORDER),
            ("LINEABOVE",     (0, 0), (-1,  0), 2,   TEAL),
            ("TOPPADDING",    (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ]))
        story.append(_sb)
        story.append(Spacer(1, 4 * mm))

        # High-risk domain alert (only if present)
        if _sp_hd:
            _hd = Table(
                [[Paragraph(
                    f"<b>Urgent:</b>  {',  '.join(_sp_hd)} — "
                    f"these areas have at least one HIGH-severity finding that needs "
                    f"immediate action (see Section 8 for full details).",
                    S["body"],
                )]],
                colWidths=[CONTENT_W],
            )
            _hd.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, -1), HIGH_BG),
                ("LINEBEFORE",    (0, 0), ( 0, -1), 4,   HIGH_FG),
                ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
                ("TOPPADDING",    (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ]))
            story.append(_hd)
            story.append(Spacer(1, 4 * mm))

        # Which areas of your AWS account have the most problems?
        if _sp_hot:
            story.append(Paragraph("Which areas of your AWS account have the most problems?", S["h3"]))
            _hot_rows = []
            for h in _sp_hot:
                _ht  = h["count"]
                _hhi = h["high_count"]
                _hc  = "#B91C1C" if _hhi > 0 else "#15803D"
                _hot_rows.append([
                    h["domain"],
                    str(_ht),
                    f"<font color='{_hc}'><b>{_hhi}</b></font>",
                ])
            story.append(_tbl(
                ["Area", "Total Issues", "Urgent (HIGH)"],
                _hot_rows,
                col_widths=[6*cm, 3.5*cm, CONTENT_W - 9.5*cm],
                right_cols=[1, 2],
            ))

    story.append(PageBreak())

    # ════════════════════════════════════════════════════════════════════════
    # 8 ░░ SECURITY AUDIT
    # ════════════════════════════════════════════════════════════════════════
    story.append(_sec(
        "8.  Security Audit",
        f"Total: {len(sec_findings)}   ·   HIGH: {high_c}   ·   "
        f"MEDIUM: {med_c}   ·   LOW: {low_c}   ·   INFO: {info_c}",
    ))

    if not sec_findings:
        story.append(Paragraph("No security findings were detected.", S["body"]))
    else:
        # ── Security severity summary bar ────────────────────────────────────
        sev_bar_vals = [
            Paragraph(f"<b>{high_c}</b>",  _ps("sb_h", fontName="Helvetica-Bold", fontSize=16, textColor=HIGH_FG, alignment=TA_CENTER)),
            Paragraph(f"<b>{med_c}</b>",   _ps("sb_m", fontName="Helvetica-Bold", fontSize=16, textColor=MED_FG,  alignment=TA_CENTER)),
            Paragraph(f"<b>{low_c}</b>",   _ps("sb_l", fontName="Helvetica-Bold", fontSize=16, textColor=LOW_FG,  alignment=TA_CENTER)),
            Paragraph(f"<b>{info_c}</b>",  _ps("sb_i", fontName="Helvetica-Bold", fontSize=16, textColor=INFO_FG, alignment=TA_CENTER)),
        ]
        sev_bar_lbls = [
            Paragraph("HIGH",   _ps("sbl_h", fontName="Helvetica", fontSize=7, textColor=HIGH_FG, alignment=TA_CENTER)),
            Paragraph("MEDIUM", _ps("sbl_m", fontName="Helvetica", fontSize=7, textColor=MED_FG,  alignment=TA_CENTER)),
            Paragraph("LOW",    _ps("sbl_l", fontName="Helvetica", fontSize=7, textColor=LOW_FG,  alignment=TA_CENTER)),
            Paragraph("INFO",   _ps("sbl_i", fontName="Helvetica", fontSize=7, textColor=INFO_FG, alignment=TA_CENTER)),
        ]
        sev_bar = Table([sev_bar_vals, sev_bar_lbls], colWidths=[CONTENT_W / 4] * 4)
        sev_bar.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (0, -1), HIGH_BG),
            ("BACKGROUND",    (1, 0), (1, -1), MED_BG),
            ("BACKGROUND",    (2, 0), (2, -1), LOW_BG),
            ("BACKGROUND",    (3, 0), (3, -1), INFO_BG),
            ("BOX",           (0, 0), (-1, -1), 0.5, BORDER),
            ("INNERGRID",     (0, 0), (-1, -1), 0.3, BORDER),
            ("LINEABOVE",     (0, 0), (-1, 0),  2.5, TEAL),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(sev_bar)
        story.append(Spacer(1, 3 * mm))

        # ── AI priority groups ───────────────────────────────────────────────
        valid_prio = [p for p in ai_insights.get("security_prioritised", [])
                      if isinstance(p, dict) and p.get("priority_group")]
        if valid_prio:
            story.append(Paragraph("AI Security Remediation Plan", S["h2"]))
            story.append(_divider())
            for grp in valid_prio:
                label   = grp.get("priority_group", "")
                risk    = grp.get("business_risk", "")
                summary = grp.get("remediation_summary", "")
                count   = grp.get("findings_count", "")
                sl      = label.lower()
                col_hex = ("#B91C1C" if "critical" in sl else
                           "#B45309" if "high"     in sl else
                           "#15803D" if "medium"   in sl else "#64748B")
                story.append(Paragraph(
                    f"<b><font color='{col_hex}'>{label}</font></b>"
                    f"<font color='#64748B' size='8'>  —  {count} finding(s)</font>",
                    S["bold"],
                ))
                if risk:
                    story.append(Paragraph(f"Business risk: {risk}", S["body"]))
                if summary:
                    story.append(Paragraph(summary, S["adv"]))
                story.append(Spacer(1, 2 * mm))
            story.append(Spacer(1, 3 * mm))

        # ── Severity donut chart ─────────────────────────────────────────────
        if high_c + med_c + low_c > 0:
            try:
                donut_png = severity_donut(high_c, med_c, low_c, "Security Findings Overview")
                story.append(KeepTogether([
                    _centered_img(donut_png, width=8 * cm, height=8 * cm),
                    Paragraph("Figure 2 — Security findings by severity", S["caption"]),
                ]))
            except Exception as e:
                logger.debug(f"Security donut failed: {e}")
            story.append(Spacer(1, 3 * mm))

        # ── Findings detail table — HIGH → MEDIUM → LOW → INFO ───────────────
        # Exclude idle/unused findings — they have their own §5 section.
        story.append(Paragraph("Security Findings Detail", S["h2"]))
        ordered = (
            [f for f in sec_findings
             if f.get("severity") == "HIGH"   and f.get("category") not in ("idle_resource", "unused_resource")] +
            [f for f in sec_findings
             if f.get("severity") == "MEDIUM" and f.get("category") not in ("idle_resource", "unused_resource")] +
            [f for f in sec_findings
             if f.get("severity") == "LOW"    and f.get("category") not in ("idle_resource", "unused_resource")] +
            [f for f in sec_findings
             if f.get("severity") == "INFO"   and f.get("category") not in ("idle_resource", "unused_resource")]
        )
        sec_rows = [
            [f.get("severity",       "—"),
             f.get("domain",         "—"),
             f.get("region",         "—"),
             f.get("resource_id",    "—")[:30],
             f.get("issue",          "—"),
             f.get("recommendation", "—")]
            for f in ordered
        ]
        story.append(_tbl(
            ["Severity", "Domain", "Region", "Resource", "Issue", "Recommendation"],
            sec_rows,
            col_widths=[1.5 * cm, 2.3 * cm, 2.0 * cm, 2.8 * cm,
                        4.0 * cm, CONTENT_W - 12.6 * cm],
            sev_col=0,
        ))

    # ── Build PDF ────────────────────────────────────────────────────────────
    doc.build(story)
    logger.info(f"PDF report written to: {output_path}")


# ═══════════════════════════════════════════════════════════════════════════
# ██  E X C E L   G E N E R A T O R
# ═══════════════════════════════════════════════════════════════════════════

def generate_excel(data: dict, output_path: str) -> None:
    """
    Generate a full Excel workbook at output_path from the consolidated data dict.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # ── Style constants ────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="1A2332")
    ALT_FILL     = PatternFill("solid", fgColor="F7FAFC")
    HIGH_FILL    = PatternFill("solid", fgColor="FFF5F5")
    MED_FILL     = PatternFill("solid", fgColor="FFFAF0")
    LOW_FILL     = PatternFill("solid", fgColor="F0FFF4")
    HIGH_FONT    = Font(color="C53030", bold=True)
    MED_FONT     = Font(color="C05621", bold=True)
    LOW_FONT     = Font(color="276749", bold=True)
    WHITE_FONT   = Font(color="FFFFFF", bold=True)
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    BOLD         = Font(bold=True)
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT         = Alignment(horizontal="left", vertical="top", wrap_text=True)
    THIN         = Side(style="thin", color="E2E8F0")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    SEVERITY_STYLES = {
        "HIGH":   (HIGH_FILL,  HIGH_FONT),
        "MEDIUM": (MED_FILL,   MED_FONT),
        "LOW":    (LOW_FILL,   LOW_FONT),
        "INFO":   (None,       None),
    }

    def _write_header(ws, headers: list, row: int = 1):
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.border = BORDER
            cell.alignment = CENTER

    def _write_row(ws, row_data: list, row: int, alt: bool = False):
        fill = ALT_FILL if alt else None
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if fill:
                cell.fill = fill
            cell.border    = BORDER
            cell.alignment = LEFT

    def _auto_width(ws, extra: int = 4):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    max_len  = max(max_len, cell_len)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + extra, 60)

    def _embed_chart_image(ws, png_bytes: bytes, anchor: str):
        """Embed a matplotlib PNG into the worksheet at the given cell anchor."""
        buf = io.BytesIO(png_bytes)
        img = XLImage(buf)
        img.width  = 600
        img.height = 320
        ws.add_image(img, anchor)

    # ──────────────────────────────────────────────────────────────────────
    # Sheet 1: Summary Dashboard
    # ──────────────────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary.sheet_view.showGridLines = False

    billing    = data.get("billing", {})
    resources  = data.get("resources", [])
    recs       = data.get("recommendations", [])
    sec_findings = data.get("security_findings", [])

    total_30d   = billing.get("30d", {}).get("total", 0.0)
    total_recs  = len(recs)
    total_sav   = sum(r.get("estimated_monthly_savings_usd", 0) for r in recs)
    high_count  = sum(1 for f in sec_findings if f.get("severity") == "HIGH")
    med_count   = sum(1 for f in sec_findings if f.get("severity") == "MEDIUM")
    low_count   = sum(1 for f in sec_findings if f.get("severity") == "LOW")

    ws_summary["A1"] = "AWS Infrastructure Dashboard"
    ws_summary["A1"].font = Font(size=18, bold=True, color="1A2332")
    ws_summary["A2"] = f"Account: {data.get('account_name', 'N/A')} ({data.get('account_id', 'N/A')})"
    ws_summary["A3"] = f"Generated: {data.get('generated_at', '')}"
    ws_summary["A4"] = f"Regions: {', '.join(data.get('regions_scanned', []))}"

    kpi_headers = ["Metric", "Value"]
    kpi_rows = [
        ["30-Day Total Spend",              f"${total_30d:,.2f}"],
        ["Total Resources Scanned",          str(len(resources))],
        ["Rightsizing Recommendations",      str(total_recs)],
        ["Potential Monthly Savings",        f"${total_sav:,.2f}"],
        ["Security Findings — HIGH",         str(high_count)],
        ["Security Findings — MEDIUM",       str(med_count)],
        ["Security Findings — LOW",          str(low_count)],
        ["Regions Scanned",                  str(len(data.get("regions_scanned", [])))],
    ]
    _write_header(ws_summary, kpi_headers, row=6)
    for i, row in enumerate(kpi_rows, start=7):
        _write_row(ws_summary, row, i, alt=i % 2 == 0)

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 20

    # ──────────────────────────────────────────────────────────────────────
    # Sheet 2: Billing — Professional Cost Intelligence Dashboard
    # Three sections: Period Summary | 30-Day Service Breakdown |
    #                 Native Savings (if any)
    # ──────────────────────────────────────────────────────────────────────
    ws_billing = wb.create_sheet("Billing")
    ws_billing.sheet_view.showGridLines = False

    # ── Helper: styled section-label row (dark navy, full width) ──────────
    def _section_label(ws, label: str, row: int, ncols: int = 8):
        """Write a full-width dark section header spanning ncols columns."""
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=ncols
        )
        cell            = ws.cell(row=row, column=1, value=label)
        cell.fill       = HEADER_FILL
        cell.font       = Font(bold=True, size=10, color="FFFFFF")
        cell.alignment  = Alignment(horizontal="left", vertical="center",
                                    indent=1)
        ws.row_dimensions[row].height = 18

    # ── Helper: currency string ────────────────────────────────────────────
    def _usd(val) -> str:
        try:
            return f"${float(val):,.2f}"
        except (TypeError, ValueError):
            return "—"

    # ── Collect billing windows ────────────────────────────────────────────
    b7  = billing.get("7d",  {})
    b15 = billing.get("15d", {})
    b30 = billing.get("30d", {})

    t7  = b7.get("total",  0.0)
    t15 = b15.get("total", 0.0)
    t30 = b30.get("total", 0.0)

    svcs_30d = b30.get("by_service", [])

    # Period date ranges from the cost_analyzer period metadata
    def _period_range(b: dict) -> str:
        p = b.get("period", {})
        s = p.get("start", "")[:10]
        e = p.get("end",   "")[:10]
        return f"{s} → {e}" if s and e else "—"

    # ══════════════════════════════════════════════════════════════════════
    # BLOCK 1 — Title Banner
    # ══════════════════════════════════════════════════════════════════════
    ws_billing.merge_cells("A1:H1")
    title = ws_billing["A1"]
    title.value     = "AWS Cloud Cost Intelligence Report"
    title.font      = Font(bold=True, size=14, color="FFFFFF")
    title.fill      = HEADER_FILL
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws_billing.row_dimensions[1].height = 30

    ws_billing.merge_cells("A2:H2")
    sub = ws_billing["A2"]
    sub.value = (
        f"Account: {data.get('account_name', 'N/A')} ({data.get('account_id', 'N/A')})   |   "
        f"Regions: {len(data.get('regions_scanned', []))}   |   "
        f"Generated: {data.get('generated_at', '')[:19].replace('T', ' ')} UTC"
    )
    sub.font      = Font(italic=True, size=9, color="718096")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws_billing.row_dimensions[2].height = 16

    # ══════════════════════════════════════════════════════════════════════
    # BLOCK 2 — Cost Period Summary
    # ══════════════════════════════════════════════════════════════════════
    _section_label(ws_billing, "  COST PERIOD SUMMARY", row=4)
    _write_header(ws_billing, [
        "Period", "Date Range", "Total Cost (USD)",
        "Daily Average (USD)", "No. of Services Billed",
        "Top AWS Service", "Top Service Cost", "Top Service % of Total",
    ], row=5)
    ws_billing.row_dimensions[5].height = 20

    PERIOD_ROWS = [
        ("7 Days",  b7,  t7),
        ("15 Days", b15, t15),
        ("30 Days", b30, t30),
    ]
    b_row = 6
    for label, bdata, total in PERIOD_ROWS:
        svcs      = bdata.get("by_service", [])
        top       = svcs[0] if svcs else {}
        # Use actual days_elapsed from billing data; fall back to the label
        # integer ("7 Days" → 7) only when running without live cost data.
        days      = bdata.get("days_elapsed") or int(label.split()[0])
        daily_avg = total / days if days else 0
        top_pct   = (top.get("cost", 0) / total * 100) if total else 0

        _write_row(ws_billing, [
            label,
            _period_range(bdata),
            _usd(total),
            _usd(daily_avg),
            len(svcs),
            top.get("service", "—"),
            _usd(top.get("cost", 0)),
            f"{top_pct:.1f}%",
        ], b_row, alt=b_row % 2 == 0)
        # Bold the total cost cell
        ws_billing.cell(row=b_row, column=3).font = Font(bold=True)
        b_row += 1

    # ══════════════════════════════════════════════════════════════════════
    # BLOCK 3 — 30-Day Service Cost Breakdown
    # ══════════════════════════════════════════════════════════════════════
    _section_label(ws_billing, "  30-DAY SERVICE COST BREAKDOWN", row=b_row + 1)
    _write_header(ws_billing, [
        "#", "AWS Service", "30-Day Cost (USD)", "% of Total Spend",
        "Daily Average (USD)",
    ], row=b_row + 2)
    ws_billing.row_dimensions[b_row + 2].height = 20
    svc_data_start = b_row + 3
    # Derive the actual number of days in the 30d window from billing metadata
    # so the daily average is always computed from the real period length.
    _30d_days_xl = b30.get("days_elapsed") or 30

    for idx, svc in enumerate(svcs_30d, start=1):
        svc_cost = svc.get("cost", 0)
        pct      = (svc_cost / t30 * 100) if t30 else 0
        daily    = svc_cost / _30d_days_xl if _30d_days_xl else 0
        row_i    = svc_data_start + idx - 1

        _write_row(ws_billing, [
            idx,
            svc.get("service", "—"),
            _usd(svc_cost),
            f"{pct:.1f}%",
            _usd(daily),
        ], row_i, alt=idx % 2 == 0)

        # Highlight the top 3 services with a subtle gold tint
        if idx <= 3:
            for col_i in range(1, 6):
                cell = ws_billing.cell(row=row_i, column=col_i)
                cell.fill = PatternFill("solid", fgColor="FFF8E1")
                cell.font = Font(bold=(col_i in (2, 3)))

    svc_data_end = svc_data_start + len(svcs_30d) - 1

    # Totals row
    total_row = svc_data_end + 1
    ws_billing.cell(row=total_row, column=2).value = "TOTAL"
    ws_billing.cell(row=total_row, column=2).font  = Font(bold=True)
    ws_billing.cell(row=total_row, column=3).value = _usd(t30)
    ws_billing.cell(row=total_row, column=3).font  = Font(bold=True)
    ws_billing.cell(row=total_row, column=4).value = "100.0%"
    ws_billing.cell(row=total_row, column=4).font  = Font(bold=True)
    ws_billing.cell(row=total_row, column=5).value = _usd(t30 / _30d_days_xl) if _30d_days_xl else "—"
    for c in range(1, 6):
        ws_billing.cell(row=total_row, column=c).border    = BORDER
        ws_billing.cell(row=total_row, column=c).fill      = PatternFill("solid", fgColor="E8F4FD")
        ws_billing.cell(row=total_row, column=c).alignment = LEFT

    # ══════════════════════════════════════════════════════════════════════
    # BLOCK 4 — Native AWS Savings Opportunities (Cost Explorer)
    # ══════════════════════════════════════════════════════════════════════
    native_recs = billing.get("native_recommendations", [])
    if native_recs:
        nr_start = total_row + 2
        _section_label(ws_billing, "  AWS NATIVE SAVINGS OPPORTUNITIES (Cost Explorer)",
                       row=nr_start)
        _write_header(ws_billing, [
            "Recommendation Type", "Resource / Instance ID",
            "Current Config", "Recommended Config",
            "Est. Monthly Savings (USD)", "Lookback Period", "", "",
        ], row=nr_start + 1)
        ws_billing.row_dimensions[nr_start + 1].height = 20
        nr_row = nr_start + 2
        for idx, rec in enumerate(native_recs, start=1):
            _write_row(ws_billing, [
                rec.get("recommendation_type", rec.get("type", "—")),
                rec.get("resource_id",         rec.get("instance_id", "—")),
                rec.get("current_config",      rec.get("current_type",      "—")),
                rec.get("recommended_config",  rec.get("recommended_type",  "—")),
                _usd(rec.get("estimated_monthly_savings_usd",
                             rec.get("estimated_savings", 0))),
                rec.get("lookback_period_in_days", "—"),
                "", "",
            ], nr_row, alt=idx % 2 == 0)
            nr_row += 1
        _next_row = nr_row
    else:
        _next_row = total_row + 2

    # ══════════════════════════════════════════════════════════════════════
    # BLOCK 5 — Current Calendar-Month Billing
    # Queried with DAILY granularity: every bucket is strictly within this
    # calendar month. No cross-month mixing, not even $0.01.
    # ══════════════════════════════════════════════════════════════════════
    cm               = billing.get("current_month", {})
    cm_svcs          = cm.get("by_service",    [])
    cm_total         = cm.get("total",          0.0)
    cm_month_name    = cm.get("month_name",     "Current Month")
    cm_days_in_month = cm.get("days_in_month",  30)
    cm_days_elapsed  = cm.get("days_elapsed",   0)
    cm_days_remaining= cm.get("days_remaining", 0)
    cm_start         = cm.get("period", {}).get("start", "")
    cm_end           = cm.get("period", {}).get("end",   "")

    cm_label_row = _next_row + 1
    _section_label(
        ws_billing,
        f"  {cm_month_name.upper()}  —  CURRENT MONTH BILLING"
        f"  |  {cm_days_elapsed} of {cm_days_in_month} days "
        f"completed, {cm_days_remaining} days remaining",
        row=cm_label_row,
    )

    # Informational note row
    ws_billing.merge_cells(
        start_row=cm_label_row + 1, start_column=1,
        end_row=cm_label_row + 1,   end_column=8,
    )
    cm_note = ws_billing.cell(
        row=cm_label_row + 1, column=1,
        value=(
            f"Billing period: {cm_start}  to  {cm_end} (end exclusive)   |   "
            f"Month has {cm_days_in_month} days   |   "
            f"Granularity: Daily   |   "
            "Today's accruing charges are excluded — only finalised daily costs are shown"
        ),
    )
    cm_note.font      = Font(italic=True, size=8, color="718096")
    cm_note.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_billing.row_dimensions[cm_label_row + 1].height = 14

    _write_header(ws_billing, [
        "#", "AWS Service",
        f"{cm_month_name} Cost (USD)",
        "% of Month Total",
        f"Daily Avg (over {cm_days_elapsed} days)",
    ], row=cm_label_row + 2)
    ws_billing.row_dimensions[cm_label_row + 2].height = 20

    cm_data_start = cm_label_row + 3
    if cm_svcs:
        for idx, svc in enumerate(cm_svcs, start=1):
            svc_cost = svc.get("cost", 0)
            pct      = (svc_cost / cm_total * 100) if cm_total else 0
            daily    = (svc_cost / cm_days_elapsed) if cm_days_elapsed else 0
            row_i    = cm_data_start + idx - 1

            _write_row(ws_billing, [
                idx,
                svc.get("service", "—"),
                _usd(svc_cost),
                f"{pct:.1f}%",
                _usd(daily),
            ], row_i, alt=idx % 2 == 0)

            # Top-3 gold highlight
            if idx <= 3:
                for col_i in range(1, 6):
                    cell = ws_billing.cell(row=row_i, column=col_i)
                    cell.fill = PatternFill("solid", fgColor="FFF8E1")
                    cell.font = Font(bold=(col_i in (2, 3)))

        cm_total_row = cm_data_start + len(cm_svcs)
    else:
        ws_billing.merge_cells(
            start_row=cm_data_start, start_column=1,
            end_row=cm_data_start,   end_column=5,
        )
        no_data = ws_billing.cell(
            row=cm_data_start, column=1,
            value="No finalised billing data available yet for this month.",
        )
        no_data.font      = Font(italic=True, color="718096")
        no_data.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cm_total_row = cm_data_start + 1

    # Month-to-date totals row
    ws_billing.cell(row=cm_total_row, column=2).value = "MONTH-TO-DATE TOTAL"
    ws_billing.cell(row=cm_total_row, column=2).font  = Font(bold=True)
    ws_billing.cell(row=cm_total_row, column=3).value = _usd(cm_total)
    ws_billing.cell(row=cm_total_row, column=3).font  = Font(bold=True)
    ws_billing.cell(row=cm_total_row, column=4).value = "100.0%" if cm_svcs else ""
    ws_billing.cell(row=cm_total_row, column=4).font  = Font(bold=True)
    ws_billing.cell(row=cm_total_row, column=5).value = (
        _usd(cm_total / cm_days_elapsed) if cm_days_elapsed else "—"
    )
    for c in range(1, 6):
        ws_billing.cell(row=cm_total_row, column=c).border    = BORDER
        ws_billing.cell(row=cm_total_row, column=c).fill      = PatternFill("solid", fgColor="E8F4FD")
        ws_billing.cell(row=cm_total_row, column=c).alignment = LEFT

    # ── Column widths (fixed, tuned to content) ────────────────────────────
    # A=# / Period, B=Service/Date Range, C=Cost, D=%/Services, E=Daily avg
    # F=Top Service name, G=Top Service cost, H=Top Service % (table 1 only)
    for col_letter, width in [
        ("A", 10), ("B", 36), ("C", 22), ("D", 22),
        ("E", 26), ("F", 32), ("G", 20), ("H", 22),
    ]:
        ws_billing.column_dimensions[col_letter].width = width

    # ──────────────────────────────────────────────────────────────────────
    # Sheet 3: All Resources  (sorted by service, grouped with dividers)
    # Columns: #  Service  Name  Region  Status  Config Detail  Tags
    # ARN and Type are omitted — too long to be useful in a spreadsheet view.
    # ──────────────────────────────────────────────────────────────────────
    ws_res = wb.create_sheet("Resources")
    ws_res.sheet_view.showGridLines = False

    res_headers = ["#", "Service", "Name", "Region", "Status", "Config Detail", "Tags"]
    _write_header(ws_res, res_headers)

    # Freeze the header row so it stays visible while scrolling
    ws_res.freeze_panes = "A2"

    # Sort resources alphabetically by service label so all S3s, EC2s, etc.
    # appear together.  Within each service, sort by region then name.
    sorted_resources = sorted(
        resources,
        key=lambda r: (
            r.get("service", "zzz").lower(),
            r.get("region", "").lower(),
            r.get("name", "").lower(),
        ),
    )

    # Style for service-group divider rows
    GROUP_FILL = PatternFill("solid", fgColor="2D3748")   # dark slate
    GROUP_FONT = Font(bold=True, color="FFFFFF", size=9)
    GROUP_ALIGN = Alignment(horizontal="left", vertical="center", indent=1)

    res_row     = 2
    seq_no      = 0
    prev_service = None

    for r in sorted_resources:
        svc    = r.get("service", "other").upper()
        meta   = r.get("metadata", {})
        status = (
            meta.get("state") or meta.get("status") or
            meta.get("cluster_status") or ""
        )
        config_detail = (
            meta.get("instance_type") or meta.get("instance_class") or
            meta.get("runtime")        or meta.get("engine") or
            meta.get("node_type")      or meta.get("volume_type") or
            meta.get("lb_type")        or meta.get("document_type") or
            meta.get("stream_mode")    or meta.get("type") or""
        )
        tags_items = list((r.get("tags") or {}).items())
        tags_str   = "; ".join(f"{k}={v}" for k, v in tags_items[:4])[:80]

        # Insert a group-divider row whenever the service changes
        if svc != prev_service:
            # Merge columns A-G for the divider label
            ws_res.merge_cells(
                start_row=res_row, start_column=1,
                end_row=res_row,   end_column=len(res_headers),
            )
            div_cell            = ws_res.cell(row=res_row, column=1,
                                              value=f"  {svc}")
            div_cell.fill       = GROUP_FILL
            div_cell.font       = GROUP_FONT
            div_cell.alignment  = GROUP_ALIGN
            # Apply the same fill to all merged cells' borders
            for col_idx in range(1, len(res_headers) + 1):
                c = ws_res.cell(row=res_row, column=col_idx)
                c.fill   = GROUP_FILL
                c.border = BORDER
            res_row    += 1
            seq_no      = 0       # reset per-group sequence
            prev_service = svc

        seq_no += 1
        alt     = seq_no % 2 == 0
        _write_row(ws_res, [
            seq_no,
            svc,
            r.get("name", ""),
            r.get("region", ""),
            str(status)[:30],
            str(config_detail)[:40],
            tags_str,
        ], res_row, alt=alt)
        res_row += 1

    # Fixed column widths — more readable than _auto_width for a wide dataset
    col_widths_res = {"A": 6, "B": 18, "C": 40, "D": 16,
                      "E": 14, "F": 22, "G": 38}
    for col_ltr, width in col_widths_res.items():
        ws_res.column_dimensions[col_ltr].width = width

    # Centre the # column
    for row_cells in ws_res.iter_rows(min_row=2, max_row=res_row - 1,
                                      min_col=1, max_col=1):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ──────────────────────────────────────────────────────────────────────
    # Sheet 4: Recommendations
    # AI guidance (if available) is injected as an extra column per resource
    # ──────────────────────────────────────────────────────────────────────
    ws_recs = wb.create_sheet("Recommendations")

    # Build a lookup: resource_id -> single-line AI guidance
    _ai_insights_obj   = data.get("ai_insights", {})
    _recs_enhanced     = _ai_insights_obj.get("recommendations_enhanced", [])
    _ai_rec_lookup: dict = {}
    for _er in (_recs_enhanced if isinstance(_recs_enhanced, list) else []):
        if isinstance(_er, dict) and _er.get("resource_id"):
            # Use only the first implementation step as a compact one-liner.
            # Format: "<Action verb> <resource> from X to Y."
            steps = _er.get("implementation_steps", [])
            if steps:
                first_step = str(steps[0]).strip()
                # Capitalise and ensure it ends with a period
                first_step = first_step[:1].upper() + first_step[1:]
                if not first_step.endswith("."):
                    first_step += "."
                _ai_text = first_step[:120]
            elif _er.get("raw_analysis"):
                # Fallback: trim raw analysis to first sentence
                raw = str(_er["raw_analysis"]).strip()
                period_pos = raw.find(". ")
                _ai_text = (raw[:period_pos + 1] if period_pos > 0 else raw[:120])
            else:
                _ai_text = ""
            _ai_rec_lookup[_er["resource_id"]] = _ai_text

    rec_headers = ["Resource ID", "Service", "Region", "Type",
                   "Current Config", "Recommended Config", "Monthly Impact (USD)",
                   "Severity", "Reason", "Source", "AI Guidance"]
    _write_header(ws_recs, rec_headers)
    for i, rec in enumerate(recs, start=2):
        sev      = rec.get("severity", "LOW")
        rec_type = rec.get("recommendation_type", "Over-provisioned")
        rid      = rec.get("resource_name", rec.get("resource_id", ""))
        ai_text      = _ai_rec_lookup.get(rid, _ai_rec_lookup.get(rec.get("resource_id", ""), ""))
        current_cfg  = rec.get("current_config", "")
        rec_cfg      = rec.get("recommended_config", "")
        # Prepend a concise sizing action so the cell immediately shows what
        # needs to change, followed by the AI-generated implementation step.
        if rec_type == "Over-provisioned" and rec_cfg and rec_cfg != "No smaller resource available":
            _sizing_hint = f"Downsize from {current_cfg} to {rec_cfg}."
        elif rec_type == "Over-provisioned" and (not rec_cfg or rec_cfg == "No smaller resource available"):
            _sizing_hint = f"No smaller resource available for {current_cfg}."
        elif rec_type == "Under-provisioned" and rec_cfg:
            _sizing_hint = f"Upsize from {current_cfg} to {rec_cfg}."
        elif rec_type == "Under-provisioned":
            _sizing_hint = f"Upsize {current_cfg}; no specific target confirmed — review manually."
        else:
            _sizing_hint = ""
        if _sizing_hint:
            ai_text = f"{_sizing_hint} {ai_text}".strip() if ai_text else _sizing_hint
        savings  = round(rec.get("estimated_monthly_savings_usd", 0), 2)
        cost_inc = round(rec.get("estimated_monthly_cost_increase_usd", 0), 2)
        # Positive value = savings (over-provisioned).
        # Negative value = additional monthly investment required (under-provisioned).
        if rec_type == "Under-provisioned" and cost_inc:
            monthly_impact = -cost_inc
        else:
            monthly_impact = savings
        _write_row(ws_recs, [
            rid,
            rec.get("service", "").upper(),
            rec.get("region", ""),
            rec_type,
            rec.get("current_config", ""),
            rec.get("recommended_config", ""),
            monthly_impact,
            sev,
            rec.get("reason", "")[:200],
            rec.get("source", ""),
            ai_text,
        ], i, alt=i % 2 == 0)
        # Colour the Type cell: under-provisioned = orange, over = yellow, review = grey
        type_cell = ws_recs.cell(row=i, column=4)
        if rec_type == "Under-provisioned":
            type_cell.fill = PatternFill("solid", fgColor="FFE0CC")
            type_cell.font = Font(color="C05621", bold=True)
        elif rec_type == "Over-provisioned":
            type_cell.fill = PatternFill("solid", fgColor="FFF9C4")
            type_cell.font = Font(color="8B7000", bold=True)
        # Colour severity cell (col 8 after adding Type)
        sev_cell = ws_recs.cell(row=i, column=8)
        fill, font = SEVERITY_STYLES.get(sev, (None, None))
        if fill:
            sev_cell.fill = fill
        if font:
            sev_cell.font = font
    _auto_width(ws_recs)
    # AI Guidance and Reason columns — allow a bit more width but keep compact
    ws_recs.column_dimensions[get_column_letter(11)].width = 55
    ws_recs.column_dimensions[get_column_letter(9)].width  = 45

    # ──────────────────────────────────────────────────────────────────────
    # Sheet 5: Security Findings
    # Redesigned: summary scorecard at top + filtered/grouped data table below
    # ──────────────────────────────────────────────────────────────────────
    ws_sec = wb.create_sheet("Security Findings")
    ws_sec.sheet_view.showGridLines = False

    # ── Top-of-sheet summary: domain × severity breakdown ─────────────────
    ws_sec.cell(row=1, column=1,
                value="Findings by Domain").font = Font(bold=True, size=11, color="1A2332")

    _sec_summary_start = 1

    _sec_domain_counts: dict = {}
    for _f in sec_findings:
        if _f.get("severity") == "INFO":
            continue   # INFO findings excluded — not actionable for most teams
        _dom = _f.get("domain", "Other")
        _sev = _f.get("severity", "LOW")
        if _dom not in _sec_domain_counts:
            _sec_domain_counts[_dom] = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
        _sec_domain_counts[_dom][_sev] = _sec_domain_counts[_dom].get(_sev, 0) + 1

    _write_header(ws_sec, ["Domain", "HIGH", "MEDIUM", "LOW", "Total"], row=2)
    _sum_row = 3
    for _dom in sorted(_sec_domain_counts.keys()):
        _dc = _sec_domain_counts[_dom]
        _h, _m, _l = _dc.get("HIGH", 0), _dc.get("MEDIUM", 0), _dc.get("LOW", 0)
        _write_row(ws_sec, [_dom, _h, _m, _l, _h + _m + _l], _sum_row,
                   alt=_sum_row % 2 == 0)
        _sum_row += 1

    # Totals footer row
    _all_h = sum(1 for _f in sec_findings if _f.get("severity") == "HIGH")
    _all_m = sum(1 for _f in sec_findings if _f.get("severity") == "MEDIUM")
    _all_l = sum(1 for _f in sec_findings if _f.get("severity") == "LOW")
    _write_row(ws_sec, ["TOTAL", _all_h, _all_m, _all_l, _all_h + _all_m + _all_l],
               _sum_row)
    for _c in range(1, 6):
        ws_sec.cell(row=_sum_row, column=_c).font = BOLD

    # ── Actionable findings table (HIGH / MEDIUM / LOW only) ─────────────
    DATA_HDR_ROW = _sum_row + 2   # One blank row gap after the summary
    sec_headers  = ["#", "Severity", "Domain", "Region",
                    "Resource ID", "Resource Type", "Issue", "Recommendation"]
    _write_header(ws_sec, sec_headers, row=DATA_HDR_ROW)

    _displayable = [_f for _f in sec_findings if _f.get("severity", "INFO") != "INFO"]
    _data_row    = DATA_HDR_ROW + 1
    _prev_sev    = None
    for _idx, _f in enumerate(_displayable, start=1):
        _sev = _f.get("severity", "LOW")

        # Insert a blank separator row between severity groups (HIGH → MEDIUM → LOW)
        if _prev_sev is not None and _sev != _prev_sev:
            _data_row += 1
        _prev_sev = _sev

        _row_data = [
            _idx,
            _sev,
            _f.get("domain", ""),
            _f.get("region", ""),
            _f.get("resource_id", ""),
            _f.get("resource_type", ""),
            _f.get("issue", "")[:300],
            _f.get("recommendation", "")[:300],
        ]
        _fill, _font = SEVERITY_STYLES.get(_sev, (None, None))
        for _col, _val in enumerate(_row_data, start=1):
            _cell           = ws_sec.cell(row=_data_row, column=_col, value=_val)
            _cell.border    = BORDER
            _cell.alignment = LEFT
            if _fill:
                _cell.fill = _fill   # Shade entire row for quick visual scanning

        # Bold-colour the severity label cell specifically
        if _font:
            ws_sec.cell(row=_data_row, column=2).font = _font

        ws_sec.row_dimensions[_data_row].height = 42   # Room for wrapped text
        _data_row += 1

    # Fixed column widths tuned to content — prevents narrow/unreadable columns
    for _col_let, _w in [("A", 5), ("B", 10), ("C", 14), ("D", 16),
                          ("E", 30), ("F", 32), ("G", 55), ("H", 55)]:
        ws_sec.column_dimensions[_col_let].width = _w

    # ──────────────────────────────────────────────────────────────────────
    # Sheet 6: AI Insights — executive summary
    # ──────────────────────────────────────────────────────────────────────
    exec_summary = _ai_insights_obj.get("executive_summary", "")
    if exec_summary:
        ws_ai = wb.create_sheet("AI Insights")
        ws_ai.sheet_view.showGridLines = False

        # Title block
        ws_ai.merge_cells("A1:F1")
        title_cell = ws_ai["A1"]
        title_cell.value = "AI-Generated Executive Summary"
        title_cell.font      = Font(bold=True, size=15, color="FFFFFF")
        title_cell.fill      = HEADER_FILL
        title_cell.alignment = CENTER
        ws_ai.row_dimensions[1].height = 28

        ws_ai.merge_cells("A2:F2")
        meta_cell = ws_ai["A2"]
        meta_cell.value = (
            f"Account: {data.get('account_name', 'N/A')}  |  "
            f"Generated: {data.get('generated_at', '')[:19].replace('T', ' ')} UTC  |  "
            f"Model: Gemini 2.5 Pro"
        )
        meta_cell.font      = Font(italic=True, size=9, color="718096")
        meta_cell.alignment = CENTER
        ws_ai.row_dimensions[2].height = 16

        ws_ai.merge_cells("A4:F4")
        body_cell = ws_ai["A4"]
        body_cell.value     = exec_summary
        body_cell.font      = Font(size=10, color="1A2332")
        body_cell.alignment = Alignment(wrap_text=True, vertical="top",
                                        horizontal="left")
        ws_ai.row_dimensions[4].height = 110

        for _ltr in ("A", "B", "C", "D", "E", "F"):
            ws_ai.column_dimensions[_ltr].width = 28

    # ──────────────────────────────────────────────────────────────────────
    # Sheet 7: AI Remediation Priority
    # ──────────────────────────────────────────────────────────────────────
    _sec_prio = _ai_insights_obj.get("security_prioritised", [])

    # Normalise: Gemini occasionally uses different field names for the key
    # columns.  Remap common variants to the canonical names before filtering.
    _FIELD_ALIASES = {
        "priority_group": ("Priority Band", "group", "priority", "band", "priority_level"),
        "findings_count": ("count", "finding_count", "num_findings"),
        "business_risk":  ("risk", "business_impact", "impact"),
        "remediation_summary": ("summary", "remediation", "actions", "recommended_actions"),
        "affected_resources":  ("resources", "resource_ids", "affected"),
    }
    for _item in _sec_prio:
        if not isinstance(_item, dict):
            continue
        for _canonical, _aliases in _FIELD_ALIASES.items():
            if not _item.get(_canonical):
                for _alias in _aliases:
                    if _item.get(_alias):
                        _item[_canonical] = _item[_alias]
                        break

    _valid_prio = [
        p for p in _sec_prio
        if isinstance(p, dict) and p.get("priority_group")
    ]

    # Always create the sheet when AI was enabled, even on partial/raw results.
    if _ai_insights_obj and (not _ai_insights_obj.get("errors") or _sec_prio or
                              _ai_insights_obj.get("executive_summary")):
        ws_prio = wb.create_sheet("AI Remediation Priority")
        ws_prio.sheet_view.showGridLines = False

        # ── Header banner ────────────────────────────────────────────────
        ws_prio.merge_cells("A1:F1")
        _banner = ws_prio["A1"]
        _banner.value     = "AI Security Remediation Priority Plan"
        _banner.font      = Font(bold=True, size=15, color="FFFFFF")
        _banner.fill      = HEADER_FILL
        _banner.alignment = CENTER
        ws_prio.row_dimensions[1].height = 30

        ws_prio.merge_cells("A2:F2")
        _sub = ws_prio["A2"]
        _sub.value = (
            f"Account: {data.get('account_name', 'N/A')}  |  "
            f"Generated: {data.get('generated_at', '')[:19].replace('T', ' ')} UTC  |  "
            f"{sum(1 for f in sec_findings if f.get('severity') in ('HIGH','MEDIUM','LOW'))} "
            f"actionable findings across "
            f"{len(set(f.get('domain','') for f in sec_findings))} domains"
        )
        _sub.font      = Font(italic=True, size=9, color="718096")
        _sub.alignment = CENTER
        ws_prio.row_dimensions[2].height = 16

        # ── Priority colour map ──────────────────────────────────────────
        PRIO_STYLES = {
            "critical": (PatternFill("solid", fgColor="C0392B"), Font(color="FFFFFF", bold=True, size=9)),
            "high":     (PatternFill("solid", fgColor="E67E22"), Font(color="FFFFFF", bold=True, size=9)),
            "medium":   (PatternFill("solid", fgColor="F1C40F"), Font(color="1A2332", bold=True, size=9)),
            "low":      (PatternFill("solid", fgColor="27AE60"), Font(color="FFFFFF", bold=True, size=9)),
        }
        FIX_WINDOW = {
            "critical": "24 hours",
            "high":     "1 week",
            "medium":   "1 month",
            "low":      "next quarter",
        }

        def _prio_fill(label: str):
            for k, v in PRIO_STYLES.items():
                if k in label.lower():
                    return v
            return (ALT_FILL, BOLD)

        # ── Column headers at row 4 ──────────────────────────────────────
        prio_hdrs = [
            "Priority Band", "Fix Within", "Findings",
            "Business Risk", "Recommended Actions", "Affected Resources",
        ]
        _write_header(ws_prio, prio_hdrs, row=4)
        ws_prio.row_dimensions[4].height = 22

        _pr = 5
        if not _valid_prio:
            # AI prioritisation data unavailable — surface the raw response if
            # present so the analyst isn't left with a blank sheet.
            ws_prio.merge_cells("A5:F5")
            _no_data = ws_prio["A5"]
            _raw_text = next(
                (str(p.get("raw_analysis", "")) for p in _sec_prio
                 if isinstance(p, dict) and p.get("raw_analysis")),
                "AI prioritisation data not available for this run. "
                "Re-run with --ai to generate the remediation plan.",
            )
            _no_data.value     = _raw_text
            _no_data.font      = Font(italic=True, size=9, color="718096")
            _no_data.alignment = Alignment(wrap_text=True, vertical="top")
            ws_prio.row_dimensions[5].height = max(60, min(len(_raw_text) // 6, 240))
        else:
            for _pg in _valid_prio:
                _label    = _pg.get("priority_group", "")
                _fix      = next((v for k, v in FIX_WINDOW.items() if k in _label.lower()), "—")
                _count    = _pg.get("findings_count", "")
                # No truncation — let Excel wrap_text handle display
                _risk     = _pg.get("business_risk", "")
                _summary  = _pg.get("remediation_summary", "")
                _affected = _pg.get("affected_resources", [])
                # One resource per line, max 20
                _aff_str  = "\n".join(str(r) for r in _affected[:20])

                _pf, _ff = _prio_fill(_label)

                # Write each cell individually for full control
                for _col, _val in enumerate(
                    [_label, _fix, _count, _risk, _summary, _aff_str], start=1
                ):
                    _c = ws_prio.cell(row=_pr, column=_col, value=_val)
                    _c.border    = Border(
                        left=Side(style="thin", color="D0D7DE"),
                        right=Side(style="thin", color="D0D7DE"),
                        top=Side(style="thin", color="D0D7DE"),
                        bottom=Side(style="medium", color="B0BEC5"),
                    )
                    _c.alignment = Alignment(wrap_text=True, vertical="top",
                                             horizontal="left")
                    if _col <= 2:
                        _c.fill = _pf
                        _c.font = _ff
                    else:
                        _c.fill = ALT_FILL if _pr % 2 == 0 else PatternFill(
                            "solid", fgColor="FFFFFF")
                        _c.font = Font(size=9)

                # Dynamic row height: base 18pt per estimated line of each text
                def _est_lines(text: str, col_chars: int) -> int:
                    if not text:
                        return 1
                    lines = text.split("\n")
                    total = 0
                    for ln in lines:
                        total += max(1, (len(ln) // col_chars) + 1)
                    return total

                # Column widths in chars used for height estimate
                _lines_risk    = _est_lines(_risk,    55)   # col D ~55 chars wide
                _lines_summary = _est_lines(_summary, 65)   # col E ~65 chars wide
                _lines_aff     = _est_lines(_aff_str,  40)  # col F ~40 chars
                _max_lines     = max(_lines_risk, _lines_summary, _lines_aff, 3)
                ws_prio.row_dimensions[_pr].height = max(52, _max_lines * 14)
                _pr += 1

        # ── Column widths ────────────────────────────────────────────────
        for _cl, _w in [("A", 22), ("B", 14), ("C", 11),
                        ("D", 55), ("E", 65), ("F", 42)]:
            ws_prio.column_dimensions[_cl].width = _w

    # ──────────────────────────────────────────────────────────────────────
    # Sheet 9: Performance Metrics
    # Per-namespace CloudWatch metric averages, min, max (7d / 30d windows)
    # and a Spike Alerts sub-table.
    # ──────────────────────────────────────────────────────────────────────
    _perf_data   = data.get("performance", {})
    # Filter to registered (user-resource) namespaces only — same whitelist as the PDF section.
    from registries.metrics_registry import METRICS_REGISTRY as _METRICS_REG_XL
    _TRACKED_NS_XL  = set(_METRICS_REG_XL.keys())
    _perf_by_ns_raw = _perf_data.get("by_namespace", {})
    _perf_by_ns  = {ns: v for ns, v in _perf_by_ns_raw.items() if ns in _TRACKED_NS_XL}
    _perf_spikes = [
        s for s in _perf_data.get("spike_correlation", [])
        if s.get("namespace") in _TRACKED_NS_XL
    ]

    ws_perf = wb.create_sheet("Performance Metrics")
    ws_perf.sheet_view.showGridLines = False

    # ── Header banner ────────────────────────────────────────────────────
    ws_perf.merge_cells("A1:I1")
    _pb = ws_perf["A1"]
    _pb.value     = "CloudWatch Performance Metrics  —  7-Day & 30-Day Summary"
    _pb.font      = Font(bold=True, size=14, color="FFFFFF")
    _pb.fill      = HEADER_FILL
    _pb.alignment = CENTER
    ws_perf.row_dimensions[1].height = 28

    ws_perf.merge_cells("A2:I2")
    _pb2 = ws_perf["A2"]
    _pb2.value = (
        f"Account: {data.get('account_name', 'N/A')}  |  "
        f"Generated: {data.get('generated_at', '')[:19].replace('T', ' ')} UTC  |  "
        f"{len(_perf_by_ns)} namespace(s)  |  "
        f"{len(_perf_spikes)} anomalie(s) detected"
    )
    _pb2.font      = Font(italic=True, size=9, color="718096")
    _pb2.alignment = CENTER
    ws_perf.row_dimensions[2].height = 15

    # ── Spike / Anomaly alerts sub-table ─────────────────────────────────
    _perf_row = 4
    ws_perf.cell(row=_perf_row, column=1,
                  value="Performance Anomalies").font = Font(bold=True, size=11, color="1A2332")
    _perf_row += 1

    if _perf_spikes:
        _write_header(ws_perf,
                      ["Severity", "Namespace", "Resource ID", "Metric",
                       "Window", "Avg Value", "Z-Score", "Unit"],
                      row=_perf_row)
        _perf_row += 1
        for _sp in sorted(_perf_spikes, key=lambda x: x.get("z_score", 0), reverse=True):
            _alt = _perf_row % 2 == 0
            _sev_sp = _sp.get("severity", "LOW")
            _write_row(ws_perf, [
                _sev_sp,
                _sp.get("namespace",   ""),
                _sp.get("resource_id", ""),
                _sp.get("metric",      ""),
                _sp.get("window",      ""),
                round(_sp.get("avg",      0), 4),
                round(_sp.get("z_score",  0), 2),
                "",   # unit populated below
            ], _perf_row, alt=_alt)
            # Colour severity cell
            _sp_fill, _sp_font = SEVERITY_STYLES.get(_sev_sp, (None, None))
            if _sp_fill:
                ws_perf.cell(row=_perf_row, column=1).fill = _sp_fill
            if _sp_font:
                ws_perf.cell(row=_perf_row, column=1).font = _sp_font
            _perf_row += 1
    else:
        _no_sp = ws_perf.cell(row=_perf_row, column=1,
                               value="No performance anomalies detected.")
        _no_sp.font = Font(italic=True, size=9, color="718096")
        _perf_row += 1

    _perf_row += 1   # blank separator row

    # ── Per-namespace metric detail ───────────────────────────────────────
    ws_perf.cell(row=_perf_row, column=1,
                  value="CloudWatch Metric Detail (7d & 30d)").font = Font(
        bold=True, size=11, color="1A2332")
    _perf_row += 1

    _detail_hdrs = [
        "Namespace", "Resource ID", "Region", "Metric", "Unit",
        "Avg (7d)", "Min (7d)", "Max (7d)",
        "Avg (30d)", "Min (30d)", "Max (30d)",
    ]
    _write_header(ws_perf, _detail_hdrs, row=_perf_row)
    _perf_row += 1

    def _metric_stats(pts: list) -> tuple:
        """Return (avg, min, max) rounded to 4dp, or blanks if no data.
        Automatically converts Bytes → GB for memory/storage metrics."""
        vals = [p["value"] for p in pts if p.get("value") is not None]
        if not vals:
            return "", "", ""
        _unit = pts[0].get("unit", "") if pts else ""
        _div  = 1024 ** 3 if _unit == "Bytes" else 1
        return (
            round(sum(vals) / len(vals) / _div, 4),
            round(min(vals) / _div, 4),
            round(max(vals) / _div, 4),
        )

    for _ns_key, _ns_windows in sorted(_perf_by_ns.items()):
        # Use 7d list; fall back to 30d if unavailable
        _res_list_7d  = _ns_windows.get("7d",  [])
        _res_list_30d = _ns_windows.get("30d", [])
        # Build lookup by resource_id for 30d
        _lookup_30d: dict = {}
        for _r30 in _res_list_30d:
            _lookup_30d[_r30["resource_id"]] = _r30

        for _r7 in _res_list_7d:
            _rid7   = _r7["resource_id"]
            _reg7   = _r7.get("region", "")
            _r30ref = _lookup_30d.get(_rid7, {})

            for _met_name, _win_data7 in _r7.get("metrics", {}).items():
                _pts7  = _win_data7.get("7d",  [])
                # 30d data for same resource + metric
                _pts30 = (_r30ref.get("metrics", {})
                          .get(_met_name, {}).get("30d", []))

                _avg7, _min7, _max7  = _metric_stats(_pts7)
                _avg30, _min30, _max30 = _metric_stats(_pts30)
                _raw_unit = _pts7[0].get("unit", "") if _pts7 else (
                            _pts30[0].get("unit", "") if _pts30 else "")
                # Display GB for storage/memory metrics that CloudWatch returns as Bytes
                _unit = "GB" if _raw_unit == "Bytes" else _raw_unit

                _alt_p = _perf_row % 2 == 0
                _write_row(ws_perf, [
                    _ns_key, _rid7, _reg7, _met_name, _unit,
                    _avg7, _min7, _max7,
                    _avg30, _min30, _max30,
                ], _perf_row, alt=_alt_p)
                _perf_row += 1

    # Column widths
    for _cl, _w in [("A", 22), ("B", 28), ("C", 14), ("D", 28), ("E", 14),
                     ("F", 12), ("G", 12), ("H", 12),
                     ("I", 12), ("J", 12), ("K", 12)]:
        ws_perf.column_dimensions[_cl].width = _w

    # Right-align all numeric columns (F onward)
    for _row_cells in ws_perf.iter_rows(
            min_row=2, max_row=_perf_row,
            min_col=6, max_col=11):
        for _cell in _row_cells:
            _cell.alignment = Alignment(horizontal="right", vertical="center")

    # ──────────────────────────────────────────────────────────────────────
    # Sheet 10: Unused / Idle Resources
    # ──────────────────────────────────────────────────────────────────────
    ws_unused = wb.create_sheet("Unused Resources")
    unused_headers = [
        "Status", "Service", "Resource ID", "Region", "Details",
        "Est. Monthly Cost (USD)", "Recommendation",
    ]
    _write_header(ws_unused, unused_headers)
    unused_row = 2

    # Human-readable service name map keyed on AWS CloudFormation resource type.
    _SVC_NAME: dict = {
        "AWS::EC2::Instance":                             "Amazon EC2",
        "AWS::EC2::Volume":                               "Amazon EBS",
        "AWS::EC2::EIP":                                  "Elastic IP",
        "AWS::EC2::VPC":                                  "Amazon VPC",
        "AWS::EC2::NatGateway":                           "NAT Gateway",
        "AWS::RDS::DBInstance":                           "Amazon RDS",
        "AWS::RDS::DBCluster":                            "Amazon Aurora",
        "AWS::ElastiCache::CacheCluster":                 "Amazon ElastiCache",
        "AWS::ElastiCache::ReplicationGroup":             "ElastiCache (Replication Group)",
        "AWS::Redshift::Cluster":                         "Amazon Redshift",
        "AWS::Lambda::Function":                          "AWS Lambda",
        "AWS::ECS::Service":                              "Amazon ECS",
        "AWS::ECS::Cluster":                              "Amazon ECS (Cluster)",
        "AWS::ElasticLoadBalancingV2::LoadBalancer":       "ALB / NLB",
        "AWS::ElasticLoadBalancing::LoadBalancer":         "Classic ELB",
        "AWS::Kinesis::Stream":                           "Amazon Kinesis",
        "AWS::SQS::Queue":                                "Amazon SQS",
        "AWS::SNS::Topic":                                "Amazon SNS",
        "AWS::ApiGateway::RestApi":                       "API Gateway (REST)",
        "AWS::ApiGatewayV2::Api":                         "API Gateway (HTTP)",
        "AWS::SageMaker::NotebookInstance":               "SageMaker Notebook",
    }
    _unused_findings = [
        f for f in sec_findings
        if f.get("category") in ("unused_resource", "idle_resource")
    ]
    _unused_findings.sort(
        key=lambda f: f.get("estimated_monthly_cost_usd", 0),
        reverse=True,
    )

    for f in _unused_findings:
        cat      = f.get("category", "unused_resource")
        status   = "Idle" if cat == "idle_resource" else "Unused"
        rtype    = f.get("resource_type", "")
        svc_name = _SVC_NAME.get(rtype, rtype)  # fall back to raw type if unmapped
        cost_val = f.get("estimated_monthly_cost_usd")
        cost_str = f"${cost_val:,.2f}" if (cost_val is not None and cost_val > 0) else "$0.00"
        _write_row(ws_unused, [
            status,
            svc_name,
            f.get("resource_id",   ""),
            f.get("region",        ""),
            f.get("issue",         "")[:250],
            cost_str,
            f.get("recommendation", "")[:250],
        ], unused_row, alt=unused_row % 2 == 0)
        # Colour the Status cell: Unused = amber, Idle = blue
        status_cell = ws_unused.cell(row=unused_row, column=1)
        if status == "Idle":
            status_cell.fill = PatternFill("solid", fgColor="2196F3")
            status_cell.font = Font(bold=True, size=9, color="FFFFFF")
        else:
            status_cell.fill = PatternFill("solid", fgColor="FF9800")
            status_cell.font = Font(bold=True, size=9, color="FFFFFF")
        unused_row += 1
    _auto_width(ws_unused)

    # ══════════════════════════════════════════════════════════════════════
    # Sheet 9: Trends & Patterns
    # ══════════════════════════════════════════════════════════════════════
    _trends_data   = data.get("trends", {})
    _ct_xl         = _trends_data.get("cost_trends", {})
    _cto_xl        = _ct_xl.get("overall", {})
    _pt_xl_list    = _trends_data.get("performance_trends", [])
    _fp_xl         = _trends_data.get("fleet_patterns",     {})
    _sp_xl         = _trends_data.get("security_patterns",  {})

    # Direction-keyed colour fills and fonts
    _XL_INC_FILL  = PatternFill("solid", fgColor="FEE2E2")   # rose  — increasing
    _XL_DEC_FILL  = PatternFill("solid", fgColor="DCFCE7")   # green — decreasing
    _XL_STA_FILL  = PatternFill("solid", fgColor="EFF6FF")   # blue  — stable
    _XL_DIR_COLOR = {"INCREASING": "B91C1C", "DECREASING": "15803D", "STABLE": "1D4ED8"}
    _XL_SEV_FILL  = {
        "HIGH":   PatternFill("solid", fgColor="FEE2E2"),
        "MEDIUM": PatternFill("solid", fgColor="FEF3C7"),
        "LOW":    PatternFill("solid", fgColor="DCFCE7"),
        "INFO":   PatternFill("solid", fgColor="EFF6FF"),
    }

    ws_tr = wb.create_sheet("Trends & Patterns")
    ws_tr.sheet_view.showGridLines = False
    tr_row = 1

    def _xl_sec(ws, title, row, ncols=9):
        """Dark navy header bar spanning ncols columns."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=title)
        c.font      = Font(bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="0F1C2E")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = BORDER
        ws.row_dimensions[row].height = 20
        return row + 1

    # ── Cost Trend Overview ──────────────────────────────────────────────
    tr_row = _xl_sec(ws_tr, "COST TREND OVERVIEW", tr_row)
    _write_header(ws_tr, [
        "Direction", "7d Daily Avg", "Prior 7d Daily", "30d Daily Avg",
        "WoW Delta $", "WoW Delta %", "Projected Monthly",
        "Confidence", "",
    ], tr_row)
    tr_row += 1
    _dir_xl   = _cto_xl.get("direction", "STABLE")
    _dir_fill = _XL_INC_FILL if _dir_xl == "INCREASING" else (_XL_DEC_FILL if _dir_xl == "DECREASING" else _XL_STA_FILL)
    _dir_font = Font(bold=True, color=_XL_DIR_COLOR.get(_dir_xl, "1D4ED8"))
    _wow_d    = _cto_xl.get("wow_delta_usd",      0)
    _wow_p    = _cto_xl.get("wow_delta_pct",      0)
    _write_row(ws_tr, [
        _dir_xl,
        f"${_cto_xl.get('daily_avg_7d',        0):,.3f}",
        f"${_cto_xl.get('prior_7d_daily_avg',  0):,.3f}",
        f"${_cto_xl.get('daily_avg_30d',        0):,.3f}",
        f"+${_wow_d:,.2f}" if _wow_d >= 0 else f"-${abs(_wow_d):,.2f}",
        f"{_wow_p:+.1f}%",
        f"${_cto_xl.get('cm_projected', _cto_xl.get('projected_monthly', 0)):,.2f}",
        _cto_xl.get("projection_confidence", "LOW"),
        "",
    ], tr_row)
    for _ci in range(1, 10):
        ws_tr.cell(row=tr_row, column=_ci).fill = _dir_fill
    ws_tr.cell(row=tr_row, column=1).font = _dir_font
    tr_row += 2

    # Cost trend by service (sorted by magnitude of change — biggest movers first)
    _ct_svc_xl = _ct_xl.get("by_service", [])
    if _ct_svc_xl:
        tr_row = _xl_sec(ws_tr, "COST TREND BY SERVICE  (sorted by change magnitude)", tr_row)
        _write_header(ws_tr, [
            "Service", "30d Daily Avg", "15d Daily Avg", "7d Daily Avg",
            "Change %", "Direction", "Trend Label", "Consistent", "",
        ], tr_row)
        tr_row += 1
        for _s in _ct_svc_xl[:25]:
            _s_dir   = _s.get("direction", "STABLE")
            _s_fill  = _XL_INC_FILL if _s_dir == "INCREASING" else (_XL_DEC_FILL if _s_dir == "DECREASING" else _XL_STA_FILL)
            _s_cons  = "Yes" if _s.get("consistent_trend") else "No"
            _write_row(ws_tr, [
                _s.get("service", ""),
                f"${_s.get('daily_avg_30d', 0):,.4f}",
                f"${_s.get('daily_avg_15d', 0):,.4f}",
                f"${_s.get('daily_avg_7d',  0):,.4f}",
                f"{_s.get('change_pct', 0):+.1f}%",
                _s_dir,
                _s.get("trend_label", ""),
                _s_cons,
                "",
            ], tr_row, alt=tr_row % 2 == 0)
            ws_tr.cell(row=tr_row, column=6).fill = _s_fill
            ws_tr.cell(row=tr_row, column=6).font = Font(bold=True, color=_XL_DIR_COLOR.get(_s_dir, "1D4ED8"))
            tr_row += 1
        tr_row += 1

    # ── Performance Metric Trends ────────────────────────────────────────
    if _pt_xl_list:
        tr_row = _xl_sec(ws_tr, "PERFORMANCE METRIC TRENDS  (30d -> 7d, severity-weighted)", tr_row)
        _write_header(ws_tr, [
            "Severity", "Resource", "Service", "Metric",
            "30d Avg", "7d Avg", "Change %", "Consistent", "Interpretation",
        ], tr_row)
        tr_row += 1
        for _pt in _pt_xl_list[:60]:
            _pt_dir  = _pt.get("direction", "STABLE")
            _pt_sev  = _pt.get("severity",  "LOW")
            _pt_cons = "Yes" if _pt.get("consistent_trend") is True else ("No" if _pt.get("consistent_trend") is False else "")
            _write_row(ws_tr, [
                _pt_sev,
                _pt.get("resource_id", ""),
                _pt.get("namespace",   "").replace("AWS/", ""),
                _pt.get("metric",      ""),
                f"{_pt.get('avg_30d', 0):.3f}",
                f"{_pt.get('avg_7d',  0):.3f}",
                f"{_pt.get('change_pct', 0):+.1f}%",
                _pt_cons,
                _pt.get("interpretation", ""),
            ], tr_row, alt=tr_row % 2 == 0)
            ws_tr.cell(row=tr_row, column=1).fill = _XL_SEV_FILL.get(_pt_sev, _XL_STA_FILL)
            ws_tr.cell(row=tr_row, column=7).fill = _XL_INC_FILL if _pt_dir == "INCREASING" else _XL_DEC_FILL
            tr_row += 1
        tr_row += 1

    # ── Fleet Provisioning Patterns ──────────────────────────────────────
    tr_row = _xl_sec(ws_tr, "FLEET PROVISIONING PATTERNS", tr_row)
    _write_header(ws_tr, [
        "Category", "Count", "Savings ($)",
        "", "", "", "", "", "",
    ], tr_row)
    tr_row += 1
    _fp_total_xl = _fp_xl.get("total_resources",        0)
    _fp_over_xl  = _fp_xl.get("overprovisioned_count",  0)
    _fp_under_xl = _fp_xl.get("underprovisioned_count", 0)
    _fp_right_xl = max(_fp_total_xl - _fp_over_xl - _fp_under_xl, 0)
    _fp_fleet_rows = [
        ("Over-provisioned",  _fp_over_xl,  f"${_fp_xl.get('total_savings_potential', 0):,.2f}"),
        ("Under-provisioned", _fp_under_xl, ""),
        ("Right-sized",       _fp_right_xl, ""),
        ("Review / No Data",  _fp_xl.get("review_count", 0), ""),
        ("Total",             _fp_total_xl, ""),
    ]
    for _fcat, _fcnt, _fsav in _fp_fleet_rows:
        _write_row(ws_tr, [
            _fcat, _fcnt, _fsav, "", "", "", "", "", "",
        ], tr_row, alt=tr_row % 2 == 0)
        if _fcat == "Over-provisioned":
            ws_tr.cell(row=tr_row, column=1).fill = PatternFill("solid", fgColor="FEF3C7")
        elif _fcat == "Under-provisioned":
            ws_tr.cell(row=tr_row, column=1).fill = PatternFill("solid", fgColor="FEE2E2")
        elif _fcat == "Right-sized":
            ws_tr.cell(row=tr_row, column=1).fill = PatternFill("solid", fgColor="DCFCE7")
        tr_row += 1

    # Systemic issues
    _sys_xl = _fp_xl.get("systemic_issues", [])
    if _sys_xl:
        tr_row += 1
        ws_tr.merge_cells(start_row=tr_row, start_column=1, end_row=tr_row, end_column=9)
        _si_h = ws_tr.cell(row=tr_row, column=1, value="SYSTEMIC ISSUES")
        _si_h.font = Font(bold=True, size=9, color="7C2D12")
        _si_h.fill = PatternFill("solid", fgColor="FEF3C7")
        tr_row += 1
        for _si in _sys_xl:
            ws_tr.merge_cells(start_row=tr_row, start_column=1, end_row=tr_row, end_column=9)
            _si_c = ws_tr.cell(row=tr_row, column=1, value=f"* {_si}")
            _si_c.font      = Font(size=8.5)
            _si_c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws_tr.row_dimensions[tr_row].height = 22
            tr_row += 1
    tr_row += 1

    # ── Security Finding Patterns ────────────────────────────────────────
    _sp_sev_xl = _sp_xl.get("severity_distribution", {})
    tr_row = _xl_sec(ws_tr, "SECURITY FINDING PATTERNS", tr_row)
    # Severity summary row
    _write_header(ws_tr, [
        "HIGH", "MEDIUM", "LOW", "INFO", "TOTAL",
        "High-Risk Domains", "", "", "",
    ], tr_row)
    tr_row += 1
    _sp_hi_d = _sp_xl.get("high_risk_domains", [])
    _write_row(ws_tr, [
        _sp_sev_xl.get("HIGH",   0),
        _sp_sev_xl.get("MEDIUM", 0),
        _sp_sev_xl.get("LOW",    0),
        _sp_sev_xl.get("INFO",   0),
        _sp_xl.get("total_findings", 0),
        ", ".join(_sp_hi_d) if _sp_hi_d else "None",
        "", "", "",
    ], tr_row)
    ws_tr.cell(row=tr_row, column=1).fill = PatternFill("solid", fgColor="FEE2E2")
    ws_tr.cell(row=tr_row, column=2).fill = PatternFill("solid", fgColor="FEF3C7")
    ws_tr.cell(row=tr_row, column=3).fill = PatternFill("solid", fgColor="DCFCE7")
    tr_row += 2

    # Domain hotspot table
    _svc_hot_xl = _sp_xl.get("services_by_findings", [])
    if _svc_hot_xl:
        _write_header(ws_tr, [
            "Domain", "Total Findings", "HIGH Findings",
            "", "", "", "", "", "",
        ], tr_row)
        tr_row += 1
        for _sh in _svc_hot_xl:
            _h_tot = _sh.get("count",      0)
            _h_hi  = _sh.get("high_count", 0)
            _write_row(ws_tr, [
                _sh.get("domain", ""),
                _h_tot, _h_hi,
                "", "", "", "", "", "",
            ], tr_row, alt=tr_row % 2 == 0)
            tr_row += 1

    # Column widths for the 9-column layout
    _tr_widths = {
        "A": 32, "B": 18, "C": 18, "D": 16,
        "E": 14, "F": 16, "G": 18, "H": 12, "I": 40,
    }
    for _cl, _cw in _tr_widths.items():
        ws_tr.column_dimensions[_cl].width = _cw

    wb.save(output_path)
    logger.info(f"Excel report written to: {output_path}")


# ═══════════════════════════════════════════════════════════════════════════
# ██  DISPATCHER
# ═══════════════════════════════════════════════════════════════════════════

def generate(data: dict, output_dir: str, fmt: str = "both") -> dict:
    """
    Generate reports in the requested format(s).

    Args:
        data:       Consolidated data dict from main.py
        output_dir: Directory to write output files
        fmt:        "pdf" | "excel" | "both"

    Returns:
        { "pdf": path | None, "excel": path | None }
    """
    os.makedirs(output_dir, exist_ok=True)
    date_str   = datetime.utcnow().strftime("%Y-%m-%d_%H-%M")
    account_id = data.get("account_id", "unknown")

    output_paths = {"pdf": None, "excel": None}

    if fmt in ("pdf", "both"):
        pdf_path = os.path.join(output_dir, f"aws_dashboard_{account_id}_{date_str}.pdf")
        logger.info("Generating PDF report...")
        try:
            generate_pdf(data, pdf_path)
            output_paths["pdf"] = pdf_path
        except Exception as e:
            logger.error(f"PDF generation failed: {e}", exc_info=True)

    if fmt in ("excel", "both"):
        excel_path = os.path.join(output_dir, f"aws_dashboard_{account_id}_{date_str}.xlsx")
        logger.info("Generating Excel report...")
        try:
            generate_excel(data, excel_path)
            output_paths["excel"] = excel_path
        except Exception as e:
            logger.error(f"Excel generation failed: {e}", exc_info=True)

    return output_paths
